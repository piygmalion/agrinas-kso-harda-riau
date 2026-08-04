# -*- coding: utf-8 -*-
"""Peta jaringan Cluster Kabupaten/Kota — Konsentrasi Lahan Sitaan PKH.

Unit II Harda · Ditreskrimum Polda Riau
Output:
  - Peta_Cluster_Kabupaten_Konsentrasi_Lahan_Sitaan_PKH.png
  - Peta_Cluster_Kabupaten_Konsentrasi_Lahan_Sitaan_PKH.html
  - output/cluster_kabupaten_pkh/*.csv
"""

from __future__ import annotations

import json
import math
from collections import defaultdict
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import networkx as nx
import numpy as np
import openpyxl
from matplotlib.lines import Line2D

BASE = Path(r"C:\Users\Patron\Downloads\sawit lagi")
XLSX = BASE / "matriks_agrinas_kso_12_polres.xlsx"
OUT_DIR = BASE / "output" / "cluster_kabupaten_pkh"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# Polres → Kabupaten/Kota resmi
POLRES_TO_KAB = {
    "Inhu": "Kab. Indragiri Hulu",
    "Inhil": "Kab. Indragiri Hilir",
    "Rohil": "Kab. Rokan Hilir",
    "Rohul": "Kab. Rokan Hulu",
    "Kuansing": "Kab. Kuantan Singingi",
    "Pelalawan": "Kab. Pelalawan",
    "Bengkalis": "Kab. Bengkalis",
    "Siak": "Kab. Siak",
    "Kampar": "Kab. Kampar",
    "Dumai": "Kota Dumai",
    "Kep. Meranti": "Kab. Kep. Meranti",
    "Pekanbaru": "Kota Pekanbaru",
}

# Agregat lokal yang lebih lengkap dari laporan (dipakai jika > sum baris kebun)
# Sumber: laporan analisis Unit II Harda
AGREGAT_LOKAL_HA = {
    "Rohul": 65947.06,       # sitaan PKH lokal
    "Rohil": 27655.51,       # sitaan PKH lokal
    "Kampar": 29174.71,      # Agrinas lokal 21 estate
    "Pelalawan": 5642.65,    # disita badan usaha (bukan angka anomali)
}

# Koridor geografis untuk clustering visual
KORIDOR = {
    "Utara": ["Rohul", "Rohil", "Bengkalis", "Dumai"],
    "Selatan": ["Inhu", "Kuansing", "Inhil"],
    "Tengah": ["Kampar", "Siak", "Pelalawan"],
    "Nihil/Gap": ["Kep. Meranti", "Pekanbaru"],
}

KORIDOR_COLOR = {
    "Utara": "#8B1E1E",
    "Selatan": "#1A5F4A",
    "Tengah": "#C45C26",
    "Nihil/Gap": "#5A6672",
}

BAND_HA = [
    (50000, "SANGAT TINGGI", "#6B0F0F"),
    (20000, "TINGGI", "#B32D2D"),
    (5000, "SEDANG", "#C45C26"),
    (1000, "RENDAH-SEDANG", "#B88A3D"),
    (0, "RENDAH/NIHIL", "#5A6672"),
]


def band_ha(ha: float) -> tuple[str, str]:
    for thr, name, col in BAND_HA:
        if ha >= thr:
            return name, col
    return "RENDAH/NIHIL", "#5A6672"


def load_kebun():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["kebun"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    return [dict(zip(hdr, r)) for r in rows[1:] if any(c is not None for c in r)]


def load_ringkasan():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["ringkasan_polres"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    return {r[0]: dict(zip(hdr, r)) for r in rows[1:] if r[0]}


def build_aggregates(kebun, ringkasan):
    agg = {}
    for pol in POLRES_TO_KAB:
        agg[pol] = {
            "polres": pol,
            "kabupaten": POLRES_TO_KAB[pol],
            "jml_kebun_baris": 0,
            "jml_kebun_intelkam": int(ringkasan.get(pol, {}).get("jml_kebun_intelkam") or 0),
            "ha_baris": 0.0,
            "ha_known": 0,
            "merah": int(ringkasan.get(pol, {}).get("merah") or 0),
            "kuning": int(ringkasan.get(pol, {}).get("kuning") or 0),
            "hijau": int(ringkasan.get(pol, {}).get("hijau") or 0),
            "estates": [],
        }

    for r in kebun:
        pol = (r.get("polres") or "").strip()
        if pol not in agg:
            continue
        agg[pol]["jml_kebun_baris"] += 1
        ha = r.get("luas_sita_ha")
        try:
            if ha is not None and str(ha).strip() != "":
                hav = float(ha)
                # filter anomali ekstrem (mis. angka jutaan dari Pelalawan corrupt)
                if 0 < hav < 200000:
                    agg[pol]["ha_baris"] += hav
                    agg[pol]["ha_known"] += 1
        except Exception:
            pass
        estate = (r.get("eks_perusahaan") or "").strip()
        if estate:
            agg[pol]["estates"].append(
                {
                    "estate": estate,
                    "ha": float(ha) if ha not in (None, "") and float(ha) < 200000 else None,
                    "klaster": r.get("klaster") or "",
                    "penerima_kso": r.get("penerima_kso") or "",
                    "lokasi": r.get("lokasi") or "",
                }
            )

    for pol, d in agg.items():
        lokal = AGREGAT_LOKAL_HA.get(pol)
        # Ambil yang lebih representatif: agregat lokal jika ada & lebih besar/lebih lengkap
        if lokal is not None:
            d["ha_konsentrasi"] = float(lokal)
            d["ha_sumber"] = "agregat_lokal_satker"
        else:
            d["ha_konsentrasi"] = d["ha_baris"]
            d["ha_sumber"] = "sum_baris_kebun"
        d["band"], d["color"] = band_ha(d["ha_konsentrasi"])
        for kname, pols in KORIDOR.items():
            if pol in pols:
                d["koridor"] = kname
                break
        else:
            d["koridor"] = "Lain"
    return agg


def build_graph(agg):
    """Graf: Polda → Koridor → Kabupaten → Estate (top per kab)."""
    G = nx.Graph()
    G.add_node("POLDA RIAU", ntype="polda", label="POLDA RIAU\n(Satgas PKH)", ha=0)

    for kor, pols in KORIDOR.items():
        total = sum(agg[p]["ha_konsentrasi"] for p in pols if p in agg)
        kebun = sum(agg[p]["jml_kebun_intelkam"] for p in pols if p in agg)
        G.add_node(
            kor,
            ntype="koridor",
            label=f"{kor}\n{total:,.0f} Ha",
            ha=total,
            kebun=kebun,
            color=KORIDOR_COLOR[kor],
        )
        G.add_edge("POLDA RIAU", kor, relasi="wilayah", layer="polda-koridor")

        for pol in pols:
            d = agg[pol]
            kid = d["kabupaten"]
            G.add_node(
                kid,
                ntype="kabupaten",
                label=f"{kid}\n{d['ha_konsentrasi']:,.0f} Ha",
                polres=pol,
                ha=d["ha_konsentrasi"],
                kebun=d["jml_kebun_intelkam"],
                band=d["band"],
                color=d["color"],
                koridor=kor,
                merah=d["merah"],
                kuning=d["kuning"],
                hijau=d["hijau"],
            )
            G.add_edge(kor, kid, relasi="klaster_koridor", layer="koridor-kab")

            # top estates by Ha (max 5) untuk detail jaringan
            estates = sorted(
                [e for e in d["estates"] if e.get("ha")],
                key=lambda x: x["ha"] or 0,
                reverse=True,
            )[:5]
            # jika tidak ada ha, ambil 3 nama pertama
            if not estates:
                estates = d["estates"][:3]
            for e in estates:
                eid = f"{pol}::{e['estate']}"
                kl = (e.get("klaster") or "").lower()
                if "merah" in kl:
                    ecol = "#B32D2D"
                elif "kuning" in kl:
                    ecol = "#C48A14"
                elif "hijau" in kl:
                    ecol = "#2E7D4F"
                else:
                    ecol = "#7A8792"
                ha_e = e.get("ha") or 0
                G.add_node(
                    eid,
                    ntype="estate",
                    label=(e["estate"][:28] + ("…" if len(e["estate"]) > 28 else "")),
                    ha=ha_e,
                    color=ecol,
                    polres=pol,
                    klaster=e.get("klaster") or "",
                )
                G.add_edge(kid, eid, relasi="lahan_sitaan", layer="kab-estate")
    return G


def layout_positions(G):
    """Radial: Polda center → koridor → kabupaten → estate."""
    pos = {"POLDA RIAU": (0.0, 0.0)}
    koridor_order = ["Utara", "Selatan", "Tengah", "Nihil/Gap"]
    for i, kor in enumerate(koridor_order):
        ang = -math.pi / 2 + i * (2 * math.pi / 4)
        pos[kor] = (2.2 * math.cos(ang), 2.2 * math.sin(ang))

    # kabupaten around their koridor
    for kor in koridor_order:
        kabs = [n for n, d in G.nodes(data=True) if d.get("ntype") == "kabupaten" and d.get("koridor") == kor]
        kx, ky = pos[kor]
        base_ang = math.atan2(ky, kx)
        for j, kab in enumerate(kabs):
            spread = (j - (len(kabs) - 1) / 2) * 0.55
            ang = base_ang + spread
            r = 4.0
            pos[kab] = (r * math.cos(ang), r * math.sin(ang))

            estates = [n for n in G.neighbors(kab) if G.nodes[n].get("ntype") == "estate"]
            for k, est in enumerate(estates):
                e_spread = (k - (len(estates) - 1) / 2) * 0.28
                ang_e = ang + e_spread
                pos[est] = ((r + 1.6) * math.cos(ang_e), (r + 1.6) * math.sin(ang_e))

    pos = nx.spring_layout(G, pos=pos, seed=42, k=0.9, iterations=40)
    return pos


def draw_map(G, agg, path: Path):
    pos = layout_positions(G)
    fig, ax = plt.subplots(figsize=(18, 14), facecolor="#F4F6F8")
    ax.set_facecolor("#F4F6F8")

    # edges by layer
    for layer, style, alpha, w in [
        ("polda-koridor", "solid", 0.45, 2.2),
        ("koridor-kab", "solid", 0.4, 1.6),
        ("kab-estate", "dashed", 0.28, 0.9),
    ]:
        el = [(u, v) for u, v, d in G.edges(data=True) if d.get("layer") == layer]
        nx.draw_networkx_edges(
            G, pos, edgelist=el, ax=ax, style=style, alpha=alpha, width=w, edge_color="#5A6672"
        )

    # nodes
    # polda
    nx.draw_networkx_nodes(
        G, pos, nodelist=["POLDA RIAU"], ax=ax,
        node_color="#0F2A44", node_size=4200, edgecolors="white", linewidths=2,
    )
    # koridor
    kors = [n for n, d in G.nodes(data=True) if d.get("ntype") == "koridor"]
    nx.draw_networkx_nodes(
        G, pos, nodelist=kors, ax=ax,
        node_color=[G.nodes[n]["color"] for n in kors],
        node_size=[1800 + G.nodes[n]["ha"] / 80 for n in kors],
        edgecolors="white", linewidths=1.5, alpha=0.95,
    )
    # kabupaten — size by Ha
    kabs = [n for n, d in G.nodes(data=True) if d.get("ntype") == "kabupaten"]
    sizes = []
    for n in kabs:
        ha = G.nodes[n]["ha"]
        sizes.append(900 + min(ha, 70000) / 18)
    nx.draw_networkx_nodes(
        G, pos, nodelist=kabs, ax=ax,
        node_color=[G.nodes[n]["color"] for n in kabs],
        node_size=sizes, edgecolors="white", linewidths=1.4, alpha=0.95,
    )
    # estates
    estates = [n for n, d in G.nodes(data=True) if d.get("ntype") == "estate"]
    nx.draw_networkx_nodes(
        G, pos, nodelist=estates, ax=ax,
        node_color=[G.nodes[n].get("color", "#7A8792") for n in estates],
        node_size=[220 + min(G.nodes[n].get("ha") or 0, 8000) / 25 for n in estates],
        edgecolors="white", linewidths=0.6, alpha=0.85, node_shape="s",
    )

    # labels — polda, koridor, kabupaten; estate hanya jika Ha besar
    labels = {}
    for n, d in G.nodes(data=True):
        nt = d.get("ntype")
        if nt in ("polda", "koridor", "kabupaten"):
            labels[n] = d.get("label", n)
        elif nt == "estate" and (d.get("ha") or 0) >= 4000:
            labels[n] = d.get("label", n)
    nx.draw_networkx_labels(G, pos, labels=labels, ax=ax, font_size=7.5, font_color="#1E242A")

    total_ha = sum(d["ha_konsentrasi"] for d in agg.values())
    ax.set_title(
        "Peta Jaringan Cluster Kabupaten/Kota\nKonsentrasi Lahan Sitaan PKH — Wilayah Hukum Polda Riau",
        fontsize=16, fontweight="bold", color="#0F2A44", pad=16,
    )
    ax.text(
        0.5, -0.02,
        f"Unit II Harda · Ditreskrimum Polda Riau  |  Total konsentrasi terhitung ≈ {total_ha:,.0f} Ha  |  "
        "Ukuran node ≈ luas sitaan  |  Kotak = estate (cuplikan top)  |  Agustus 2026",
        transform=ax.transAxes, ha="center", fontsize=9, color="#5A6672",
    )

    legend = [
        mpatches.Patch(color="#0F2A44", label="Polda Riau (pusat)"),
        mpatches.Patch(color=KORIDOR_COLOR["Utara"], label="Koridor Utara"),
        mpatches.Patch(color=KORIDOR_COLOR["Selatan"], label="Koridor Selatan"),
        mpatches.Patch(color=KORIDOR_COLOR["Tengah"], label="Koridor Tengah"),
        mpatches.Patch(color=KORIDOR_COLOR["Nihil/Gap"], label="Nihil / Gap-fill"),
        mpatches.Patch(color="#6B0F0F", label="Kab. konsentrasi ≥50 rb Ha"),
        mpatches.Patch(color="#B32D2D", label="Kab. konsentrasi ≥20 rb Ha"),
        mpatches.Patch(color="#C45C26", label="Kab. konsentrasi ≥5 rb Ha"),
        Line2D([0], [0], color="#5A6672", lw=1, ls="--", label="Kabupaten → estate"),
    ]
    ax.legend(handles=legend, loc="lower left", fontsize=8, framealpha=0.95, ncol=2)
    ax.axis("off")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"OK PNG: {path}")


def draw_bar_cluster(agg, path: Path):
    """Bar ranking konsentrasi Ha per kabupaten + annotation koridor."""
    rows = sorted(agg.values(), key=lambda d: d["ha_konsentrasi"], reverse=True)
    fig, ax = plt.subplots(figsize=(12, 7), facecolor="#F4F6F8")
    ax.set_facecolor("#F4F6F8")
    labels = [r["kabupaten"].replace("Kab. ", "").replace("Kota ", "") for r in rows]
    vals = [r["ha_konsentrasi"] for r in rows]
    colors = [r["color"] for r in rows]
    y = np.arange(len(rows))
    ax.barh(y, vals, color=colors, height=0.72)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=10)
    ax.invert_yaxis()
    for i, r in enumerate(rows):
        txt = f"{r['ha_konsentrasi']:,.0f} Ha · {r['jml_kebun_intelkam']} kebun · {r['koridor']}"
        ax.text(r["ha_konsentrasi"] + max(vals) * 0.01, i, txt, va="center", fontsize=8, color="#1E242A")
    ax.set_xlabel("Luas sitaan PKH (Ha)")
    ax.set_title(
        "Ranking Konsentrasi Lahan Sitaan PKH per Kabupaten/Kota\nUnit II Harda · Ditreskrimum Polda Riau",
        fontsize=13, fontweight="bold", color="#0F2A44",
    )
    ax.set_xlim(0, max(vals) * 1.35 if vals else 1)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    ax.grid(axis="x", alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=160, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"OK PNG: {path}")


def write_html(G, agg, path: Path):
    nodes, edges = [], []
    idmap = {n: i for i, n in enumerate(G.nodes())}
    for n, d in G.nodes(data=True):
        nt = d.get("ntype")
        if nt == "polda":
            size, shape, color = 40, "dot", "#0F2A44"
        elif nt == "koridor":
            size, shape, color = 28, "dot", d.get("color", "#5A6672")
        elif nt == "kabupaten":
            size = 16 + min(d.get("ha", 0), 70000) / 2500
            shape, color = "box", d.get("color", "#5A6672")
        else:
            size = 8 + min(d.get("ha") or 0, 8000) / 800
            shape, color = "diamond", d.get("color", "#7A8792")
        title = n if nt != "estate" else n.split("::", 1)[-1]
        bits = [title, f"tipe: {nt}", f"Ha: {d.get('ha', 0):,.0f}"]
        if d.get("band"):
            bits.append(f"band: {d['band']}")
        if d.get("kebun") is not None and nt == "kabupaten":
            bits.append(f"kebun Intelkam: {d['kebun']}")
            bits.append(f"M/K/H: {d.get('merah',0)}/{d.get('kuning',0)}/{d.get('hijau',0)}")
        nodes.append({
            "id": idmap[n],
            "label": (d.get("label") or n).replace("\n", " ")[:40],
            "title": " | ".join(bits),
            "color": color,
            "shape": shape,
            "size": size,
            "font": {"color": "#1E242A", "size": 12 if nt != "estate" else 10},
        })
    for u, v, d in G.edges(data=True):
        edges.append({
            "from": idmap[u], "to": idmap[v],
            "dashes": d.get("layer") == "kab-estate",
            "width": 2.2 if d.get("layer") == "polda-koridor" else 1.3,
            "color": {"color": "#6A7680"},
            "title": d.get("relasi", ""),
        })

    ranking = sorted(agg.values(), key=lambda x: -x["ha_konsentrasi"])
    hub_html = "".join(
        f'<div class="hub"><b>{r["kabupaten"]}</b><br/>'
        f'<small>{r["ha_konsentrasi"]:,.0f} Ha · {r["jml_kebun_intelkam"]} kebun · '
        f'{r["band"]} · {r["koridor"]}</small></div>'
        for r in ranking
    )

    html = f"""<!DOCTYPE html>
<html lang="id"><head>
<meta charset="utf-8"/>
<title>Cluster Kabupaten/Kota — Konsentrasi Lahan Sitaan PKH</title>
<script src="https://unpkg.com/vis-network@9.1.9/standalone/umd/vis-network.min.js"></script>
<style>
body{{margin:0;font-family:Segoe UI,Calibri,sans-serif;background:#0F2A44;color:#fff;display:grid;grid-template-rows:auto 1fr;height:100vh}}
header{{padding:14px 20px;border-bottom:3px solid #C45C26}}
h1{{margin:0;font-size:1.2rem}} .sub{{color:#B8C5D0;font-size:.82rem;margin-top:4px}}
.shell{{display:grid;grid-template-columns:1fr 320px;min-height:0}}
#net{{background:#F4F6F8}}
aside{{padding:14px;overflow:auto;background:rgba(15,42,68,.96)}}
h2{{font-size:.75rem;color:#C45C26;text-transform:uppercase;letter-spacing:.08em}}
.hub{{background:rgba(0,0,0,.22);border-left:4px solid #C45C26;padding:8px 10px;margin:6px 0;border-radius:0 8px 8px 0;font-size:12px}}
.note{{font-size:11px;color:#9AA8B5;line-height:1.45}}
</style></head><body>
<header>
  <h1>Cluster Kabupaten/Kota — Konsentrasi Lahan Sitaan PKH</h1>
  <div class="sub">Unit II Harda · Ditreskrimum Polda Riau · Polda → Koridor → Kabupaten/Kota → Estate (cuplikan)</div>
</header>
<div class="shell">
  <div id="net"></div>
  <aside>
    <h2>Ranking konsentrasi</h2>
    {hub_html}
    <h2>Cara baca</h2>
    <div class="note">
      Ukuran node ≈ luas sitaan PKH.<br/>
      Warna kabupaten = band konsentrasi.<br/>
      Empat koridor: Utara (eskalasi), Selatan (volume), Tengah (campuran), Nihil/Gap.<br/>
      Estate = cuplikan top per kab (bukan seluruh 130).
    </div>
  </aside>
</div>
<script>
const nodes=new vis.DataSet({json.dumps(nodes, ensure_ascii=False)});
const edges=new vis.DataSet({json.dumps(edges, ensure_ascii=False)});
new vis.Network(document.getElementById('net'), {{nodes, edges}}, {{
  physics:{{barnesHut:{{gravitationalConstant:-12000, springLength:140, springConstant:0.03}}, stabilization:{{iterations:160}}}},
  interaction:{{hover:true, navigationButtons:true}},
  edges:{{smooth:{{type:'continuous'}}}}
}});
</script></body></html>"""
    path.write_text(html, encoding="utf-8")
    # also embed JS data for file:// open of sibling if needed
    print(f"OK HTML: {path}")


def export_tables(agg):
    import csv

    path = OUT_DIR / "tabel_konsentrasi_kabupaten_pkh.csv"
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([
            "polres", "kabupaten_kota", "koridor", "ha_konsentrasi", "ha_sumber",
            "ha_baris_kebun", "jml_kebun_intelkam", "jml_kebun_baris",
            "merah", "kuning", "hijau", "band_konsentrasi",
        ])
        for r in sorted(agg.values(), key=lambda x: -x["ha_konsentrasi"]):
            w.writerow([
                r["polres"], r["kabupaten"], r["koridor"], f'{r["ha_konsentrasi"]:.2f}',
                r["ha_sumber"], f'{r["ha_baris"]:.2f}', r["jml_kebun_intelkam"],
                r["jml_kebun_baris"], r["merah"], r["kuning"], r["hijau"], r["band"],
            ])
    print(f"OK CSV: {path}")

    # koridor rollup
    path2 = OUT_DIR / "tabel_koridor_pkh.csv"
    with path2.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["koridor", "kabupaten", "ha_total", "kebun_intelkam", "merah", "kuning", "hijau"])
        for kor, pols in KORIDOR.items():
            ha = sum(agg[p]["ha_konsentrasi"] for p in pols)
            keb = sum(agg[p]["jml_kebun_intelkam"] for p in pols)
            m = sum(agg[p]["merah"] for p in pols)
            k = sum(agg[p]["kuning"] for p in pols)
            h = sum(agg[p]["hijau"] for p in pols)
            w.writerow([kor, "; ".join(POLRES_TO_KAB[p] for p in pols), f"{ha:.2f}", keb, m, k, h])
    print(f"OK CSV: {path2}")


def main():
    kebun = load_kebun()
    ringkasan = load_ringkasan()
    agg = build_aggregates(kebun, ringkasan)
    G = build_graph(agg)

    print("=== Konsentrasi Ha per Kabupaten/Kota ===")
    for r in sorted(agg.values(), key=lambda x: -x["ha_konsentrasi"]):
        print(
            f"{r['kabupaten']:28s} {r['ha_konsentrasi']:10,.1f} Ha  "
            f"kebun={r['jml_kebun_intelkam']:2d}  {r['band']:14s}  {r['koridor']}"
        )

    draw_map(G, agg, BASE / "Peta_Cluster_Kabupaten_Konsentrasi_Lahan_Sitaan_PKH.png")
    draw_bar_cluster(agg, BASE / "Peta_Cluster_Kabupaten_Ranking_Lahan_Sitaan_PKH.png")
    # also save copies to output
    draw_map(G, agg, OUT_DIR / "fig_cluster_jaringan_kabupaten_pkh.png")
    draw_bar_cluster(agg, OUT_DIR / "fig_cluster_ranking_kabupaten_pkh.png")
    write_html(G, agg, BASE / "Peta_Cluster_Kabupaten_Konsentrasi_Lahan_Sitaan_PKH.html")
    write_html(G, agg, OUT_DIR / "Peta_Cluster_Kabupaten_Konsentrasi_Lahan_Sitaan_PKH.html")
    export_tables(agg)

    # gexf export
    H = G.copy()
    for n, d in H.nodes(data=True):
        for k, v in list(d.items()):
            d[k] = "" if v is None else str(v)
    nx.write_gexf(H, OUT_DIR / "cluster_kabupaten_pkh.gexf")
    print(f"OK GEXF: {OUT_DIR / 'cluster_kabupaten_pkh.gexf'}")
    print(f"Graph: {G.number_of_nodes()} nodes, {G.number_of_edges()} edges")


if __name__ == "__main__":
    main()
