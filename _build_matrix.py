# -*- coding: utf-8 -*-
"""Build matriks_agrinas_kso_12_polres.xlsx from extracted structured data."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True, size=10)
thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin


def autosize(ws, max_width=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value:
                length = max(length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(12, length + 2)


# ===================== SHEET: kebun =====================
ws = wb.active
ws.title = "kebun"
kebun_headers = [
    "id", "polres", "eks_perusahaan", "penerima_kso", "luas_sita_ha", "luas_kso_ha",
    "lokasi", "status_kso", "klaster", "konflik_note", "sumber", "kepercayaan",
]
ws.append(kebun_headers)

kebun_rows = [
    # INHU (Intelkam + lokal)
    ("K-INHU-01", "Inhu", "PT Agrinas / BBU 1", "Bernas Mulya Mandiri / Osten Panjaitan", 4157, 4157, "Payarumbai, Seberida", "Sudah KSO", "Hijau", "Tuntutan plasma masyarakat (lokal)", "Intelkam+INHU DOCX", "sedang"),
    ("K-INHU-02", "Inhu", "PT Agrinas KAT 1", "Bernas Mulya Mandiri", 6535, 6535, "Kelesa/Payarumbai, Seberida", "Sudah KSO", "Hijau", "Nihil konflik Intelkam", "Intelkam+INHU", "sedang"),
    ("K-INHU-03", "Inhu", "PT Agrinas KAT 2", "Bernas Mulya Mandiri", 5123, 5123, "Kelesa, Seberida", "Sudah KSO", "Hijau", "Nihil", "Intelkam+INHU", "sedang"),
    ("K-INHU-04", "Inhu", "EX PT Palm Lestari Makmur", "Koperasi Jasa TKBM Siak Mahasakti Karya", 2144.14, 2144.14, "Penyaguhan, Batang Gansal", "Sudah KSO", "Hijau*", "Sengketa 560 Ha; panen paksa Apr 2026; bentrok; LP pencurian TBS", "INHU DOCX", "kuat"),
    ("K-INHU-05", "Inhu", "PT Tugu Palma Sumatera", "-", None, None, "Payarumbai, Seberida", "Satgas", "Hijau", "Bermasalah kelompok masyarakat", "INHU DOCX", "sedang"),
    ("K-INHU-06", "Inhu", "PKS PT BBU II", "Agus S Lubis", 9, 9, "Kuala Mulya, Kuala Cenaku", "Sudah KSO", "Hijau", "Nihil", "Intelkam+INHU", "kuat"),
    ("K-INHU-07", "Inhu", "PT SSK Sumber Sawit Kencana", "Sulastri", None, None, "Belimbing, Batang Gansal", "Sudah KSO", "Hijau", "Nihil", "INHU", "sedang"),
    ("K-INHU-08", "Inhu", "PT PAS Prima Agri Sawit Indo", "PT PAS / Gerbang Pintu Air", 1564.54, 1564.54, "Danau Rambai, Batang Gansal", "Sudah KSO", "Hijau", "Nihil; DQ kolom mungkin tertukar", "INHU", "sedang"),
    ("K-INHU-09", "Inhu", "PT Indrawan Perkasa", "PT Tiga Raja Mas (H.M. Ali)", 724.52, 724.52, "Sungai Akar, Batang Gansal", "Sudah KSO", "Kuning", "Bentrok antar pemanen; tutup akses jalan TBS", "Intelkam+INHU", "kuat"),
    ("K-INHU-10", "Inhu", "PT Banyu Bening Utama II", "Osten Panjaitan / Agus S Lubis", 4443, 4443, "Kuala Cenaku", "Sudah KSO", "Hijau", "Bermasalah lahan Desa Kuala Cenaku (lokal)", "INHU", "sedang"),
    ("K-INHU-11", "Inhu", "PT Seberida Subur", "Dikelola langsung Agrinas", 956, 956, "Siambul, Batang Gansal", "Langsung Agrinas", "Hijau", "Bermasalah Kelompok Arbain", "INHU", "sedang"),
    ("K-INHU-12", "Inhu", "PT Duta Palma Group", "Dikelola langsung Agrinas", None, None, "Penyaguhan, Batang Gansal", "Langsung Agrinas", "Hijau", "Tuntutan lahan masyarakat", "INHU", "sedang"),
    ("K-INHU-13", "Inhu", "PT BIM Mahkota Group", "Satgas PKH", 63.33, None, "Danau Rambai, Batang Gansal", "Satgas", "Hijau", "Pesangon karyawan", "INHU", "sedang"),
    ("K-INHU-14", "Inhu", "Perkebunan Pakpahan / PT Runggu", "PT Runggu", 2000, 2000, "Pesajian, Batang Peranap", "Sudah KSO", "Hijau", "Masyarakat + jalan rusak", "INHU", "sedang"),
    ("K-INHU-15", "Inhu", "Perkebunan Sinaga", "PT Runggu", 1000, None, "Pauh Ranap, Peranap", "Sudah KSO", "Hijau", "Masyarakat + jalan rusak", "INHU", "sedang"),
    ("K-INHU-16", "Inhu", "POKTAN SMB / Toton Naibaho", "CV Giza Gemilang", 331.68, 331.68, "Alim, Batang Cenaku", "Sudah KSO", "Hijau", "Nihil", "INHU", "sedang"),
    ("K-INHU-17", "Inhu", "PT SWP Sinar Widita Pamarta", "KUD Balai Jaya Sempurna", 57, 57, "Pasir Keranji, Sungai Lala", "Sudah KSO", "Hijau", "TBS sering dicuri", "INHU", "sedang"),
    ("K-INHU-18", "Inhu", "PT Selantai Argo Lestari", "-", 469.85, 469.85, "Talang Durian Cacar, Rakit Kulim", "Belum/—", "Kuning", "Konflik MHA Talang Mamak + karyawan", "Intelkam+INHU", "kuat"),
    ("K-INHU-19", "Inhu", "PT Tunggal Perkasa Plantation", "Poktan JD Karya Mandiri", 574, 574, "Lirik / Redang Seko", "Sudah KSO", "Hijau", "Aksi PKN Jun 2026; minta kelola bersama", "INHU", "kuat"),
    ("K-INHU-20", "Inhu", "PT Tasma Puja", "Agrinas", None, None, "Kepayang Sari, Batang Cenaku", "Agrinas", "Hijau", "Tolak jika 17,5% hasil tidak diberikan; luas OCR ambigu", "INHU", "tipis"),
    ("K-INHU-21", "Inhu", "PT Pasir Mas Giri Raya", "-", 54.33, None, "Sei Lala / Lubuk Batu Jaya", "—", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-INHU-22", "Inhu", "PT Mitra Kembang Selaras", "-", 2068, None, "Lirik", "—", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-INHU-23", "Inhu", "PT Gandaerah Hendana", "-", 118, None, "Banjar Balam / Lirik", "—", "Hijau", "Isu Redang Seko", "INHU", "sedang"),
    ("K-INHU-24", "Inhu", "PT Tesso Indah", "Dikembalikan Satgas", 2000, None, "Rengat Barat", "Dikembalikan", "Hijau", "Permasalahan karyawan", "INHU", "sedang"),
    ("K-INHU-25", "Inhu", "KKPA Tani Bahagia", "-", 645, None, "Lubuk Batu Tinggal, LBJ", "—", "Hijau", "Penolakan Hasbullah; potensi", "INHU", "sedang"),
    ("K-INHU-26", "Inhu", "PT Sumatera Makmur Lestari + KUD", "TDE Tuah Daleh Esa", 2349.23, None, "Seberida / Batang Gansal / Cenaku", "KSO", "Hijau*", "Sita plasma 355,31 Ha; penolakan", "INHU", "kuat"),
    ("K-INHU-27", "Inhu", "Masyarakat Sungai Ubo / Air Dingin", "-", 400, None, "Peranap / Batang Peranap", "Masyarakat", "Hijau", "Plang PKH hilang; tetap dikuasai masyarakat", "INHU", "sedang"),
    ("K-INHU-28", "Inhu", "PT Bintang Riau Sejahtera", "CV Aristama Karya Persada", 51, 51, "Baturijal Barat, Peranap", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    # ROHIL
    ("K-ROHIL-01", "Rohil", "PT Gunung Mas Raya (Rumbia I)", "PT Ujung Tanjung Sejahtera", 1800, 1800, "Rumbia I Estate / Balai Jaya area", "Sudah KSO", "Merah", "Bentrok 20 Okt 2025 vs W. Siringo; 7 luka; PAM non-BUJP Flores", "Intelkam", "kuat"),
    ("K-ROHIL-02", "Rohil", "PT Salim Ivomas Pratama", "Digjaya Nata Persada", 1008, 1008, "Sungai Dua / Balam Sempurna / Balai Jaya", "Sudah KSO", "Kuning", "Penolakan penyerahan KSO", "Intelkam", "kuat"),
    ("K-ROHIL-03", "Rohil", "PT Cibaliung Tunggal Plantation", "Koperasi Parit Nantinggi", 394.24, 394.24, "Area II Salim Ivomas, Balai Jaya", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-ROHIL-04", "Rohil", "PT Rama Salomo (Hendri Sianipar)", "Kop. Produsen Batang Kumu", 420, 420, "Simpang Kanan", "Sudah KSO", "Kuning", "Penolakan karyawan/pengelola", "Intelkam", "sedang"),
    ("K-ROHIL-05", "Rohil", "PT Rama Salomo", "Kop. Produsen Rumpun Sejahtera", 602.31, 602.31, "Tanjung Medan", "Sudah KSO", "Kuning", "Penolakan", "Intelkam", "sedang"),
    ("K-ROHIL-06", "Rohil", "PT Andika Permata Sawit Lestari", "PT Satahi / Poktan Melayu Terpadu", 3200, 3200, "Pujud / Siarang Arang", "Sudah KSO", "Kuning", "Penolakan vs Agrinas/KSO", "Intelkam", "kuat"),
    ("K-ROHIL-07", "Rohil", "PT Bumi Riau Bina Makmur", "PT BWB / Koperasi K21", 1988, None, "Tanah Putih", "Sudah KSO", "Kuning", "Penolakan plang", "Intelkam", "sedang"),
    ("K-ROHIL-08", "Rohil", "Agregat sitaan PKH Rohil", "8 KSO (nama tidak lengkap di sumber)", 27655.51, 27655.51, "Kab. Rokan Hilir", "Campuran", "1M/6K/17H", "LP Agrinas=3; pengaduan masyarakat tinggi", "Bismillah+Intelkam", "sedang"),
    # ROHUL
    ("K-ROHUL-01", "Rohul", "PT BK 1 / Berkat Satu", "PT Nusantara Sawit Majuma", 1383.92, 1383.92, "Sontang, Bonai Darussalam", "Sudah KSO", "Merah", "Bentrok PAM 7 Feb 2026; 1 MD + 6 luka; PAM non-BUJP Nias", "Intelkam+Rohul", "kuat"),
    ("K-ROHUL-02", "Rohul", "PT APSL Andika Permata", "Agrinas Palma Nusantara", 12000, 12000, "Kebun Jurong, Bonai", "Langsung Agrinas", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-ROHUL-03", "Rohul", "CV Halim Group", "Agrinas Palma Nusantara", 2344.42, 2344.42, "Desa Sontang", "Langsung Agrinas", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-ROHUL-04", "Rohul", "PT Torganda Tambusai Timur", "Penguasaan Agrinas / mitra", 12693.26, None, "Tambusai Timur", "Agrinas", "Kuning", "Pengusiran panen; PAM PP Rohul; LP pengrusakan mess", "Intelkam+Rohul", "kuat"),
    ("K-ROHUL-05", "Rohul", "PT Togos Gopas / Maju Bersama", "Penguasaan Agrinas", 1626.03, 1626.03, "Tambusai Timur", "Agrinas", "Kuning", "Pok Ahmad Lubis vs PAM Torus Ganda", "Intelkam", "kuat"),
    ("K-ROHUL-06", "Rohul", "PT Hutahean", "PT Cantya/Cintya Graha Indah", 831.43, 831.43, "Sungai Kuning, Rambah Samo", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-ROHUL-07", "Rohul", "PT Ekaudra Indonesia", "CV Ginting Sukses Abadi", 678.54, 678.54, "Kota Lama, Kunto Darussalam", "Sudah KSO", "Kuning", "Penolakan KSO dari 5 desa", "Intelkam", "kuat"),
    ("K-ROHUL-08", "Rohul", "PT Sumber Alam Makmur Sentosa", "KUD Elok Besamo", 182, 182, "Tambusai Timur", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-ROHUL-09", "Rohul", "PT Sardela", "Kop. Desa Merah Putih / PT Sardela", 197, 197, "Mura Dilam, Kunto Darussalam", "Sudah KSO", "Hijau", "Penunjukan mandiri Agrinas", "Intelkam+Rohul", "sedang"),
    ("K-ROHUL-10", "Rohul", "PT Torganda Batang Kumu 2", "Agrinas / Koperasi Sahbela", 7556.31, None, "Tambusai Utara", "Agrinas", "Kuning", "Koperasi + PAM FBI panen", "Intelkam", "sedang"),
    ("K-ROHUL-11", "Rohul", "PT Torganda Rantau Kasai", "Agrinas / RKG Sariman Siregar", 10605.32, None, "Tambusai Utara", "Agrinas", "Kuning", "Panen sepihak ~3000 Ha; LP pencurian TBS", "Intelkam+Rohul", "kuat"),
    ("K-ROHUL-12", "Rohul", "PT Torganda Karya Perdana", "Agrinas", 1374.88, None, "Tambusai Utara", "Agrinas", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-ROHUL-13", "Rohul", "KSO lain (daftar nama)", "Telangka Indo Permai; Danramil 02; Bengkalis Jangkang; Budi Murni; Kalingga; Anugrah Tuah Mulia; Duta Agrindo", None, None, "Kab. Rokan Hulu", "Sudah KSO", "Campuran", "12 KSO tercatat; mapping estate tidak lengkap", "Rohul DOCX", "tipis"),
    # INHIL
    ("K-INHIL-01", "Inhil", "PT Agro Sarimas Indonesia", "Citra Mutiara Bumi Riau", 2980.61, 2980.61, "Kempas / Bayas Jaya", "Sudah KSO", "Kuning", "Tuntutan kembalikan ke KPCH", "Intelkam", "kuat"),
    ("K-INHIL-02", "Inhil", "PT Riau Sawitindo Abadi", "CV Cipta Nugraha", 750.21, 750.21, "Belaras Barat & Batang Sari, Mandah", "Sudah KSO", "—", "Pemanenan TBS duga KSO Mei 2026; tahap penyelidikan", "Inhil konflik PDF", "kuat"),
    ("K-INHIL-03", "Inhil", "PT SAGM Tempuling", "Poktan Berkah Tani Sejahtera", 35.54, 35.54, "Tempuling", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-INHIL-04", "Inhil", "PT SAGM Batang Tuaka", "Poktan Berkah Tani Sejahtera", 124, 124, "Batang Tuaka", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-INHIL-05", "Inhil", "PT Kemuning Sawit Unggul / CV Hasil Usaha Tani", "APN / SPK", None, None, "Inhil", "KSO/APN", "Kuning", "Penghadangan masyarakat vs APN", "Intelkam", "sedang"),
    ("K-INHIL-06", "Inhil", "Kelompok Tani Naibaho", "Cahaya Puteri Melayu", 78, None, "Keritang", "—", "—", "Pencurian TBS rugi Rp21jt; bukan konflik KSO aktif", "Inhil PDF", "sedang"),
    # KUANSING
    ("K-KUANS-01", "Kuansing", "PT Cerenti Subur", "PT Agrinas Palma Nusantara", 8929, 8929, "Sikakak Cerenti; Sungai Sorik", "Belum KSO (kelola Agrinas)", "Hijau", "Banyak LP pencurian TBS 2023-2026", "Status KSO+LP", "kuat"),
    ("K-KUANS-02", "Kuansing", "PT Wana Jingga Timur", "PT Agrinas Palma Nusantara", 4196, 4196, "Cerenti / Inuman / Pesikaian", "Belum KSO (kelola Agrinas)", "Hijau", "Tuntutan masyarakat 20%; komunikasi tertutup; banyak LP", "Status KSO+Potensi", "kuat"),
    ("K-KUANS-03", "Kuansing", "PTPN IV Afd 7-9 / Pesikaian", "PT Agrinas / relokasi TNTN", 634, None, "Pesikaian, Cerenti", "Belum KSO", "Kuning", "Relokasi TNTN vs warga tempatan; Koptan Bagan Limau", "Status KSO+Potensi", "kuat"),
    ("K-KUANS-04", "Kuansing", "PT Merauke Tetap Jaya", "Wana Agri Santosa", 417.87, 417.87, "Serosa, Hulu Kuantan", "Sudah KSO", "Kuning", "Status lahan tidak jelas / isu pencurian TBS", "Intelkam", "sedang"),
    ("K-KUANS-05", "Kuansing", "PT Garuda Sakti Nusantara / Wanasari", "PT Garuda Sakti / Agrinas", 315, 315, "Sungai Buluh & Simpang Raya", "Sudah KSO", "Hijau", "Nihil", "Status KSO", "sedang"),
    ("K-KUANS-06", "Kuansing", "PT Cahaya Panam Perkasa", "Agrinas", 543, 543, "Suka Maju & Beringin Jaya", "Sudah KSO", "Hijau", "Nihil", "Status KSO", "sedang"),
    ("K-KUANS-07", "Kuansing", "PT Sailan Antau Batuah", "Agrinas", 200, 200, "Suka Maju", "Sudah KSO", "Hijau", "Nihil", "Status KSO", "sedang"),
    ("K-KUANS-08", "Kuansing", "PT Duta Palma Nusantara", "Agrinas", 14237, None, "Kopah, Kuantan Tengah", "Agrinas", "Hijau", "Nihil Intelkam", "Intelkam", "sedang"),
    ("K-KUANS-09", "Kuansing", "PT Gatipura Mulya", "CV Tiga Bintang Sinergi", 138.78, 138.78, "Sungai Langsat", "Sudah KSO", "Hijau", "Penguasaan kembali Satgas Jul 2025 + LP pencurian", "Potensi+LP", "sedang"),
    ("K-KUANS-10", "Kuansing", "PT Tri Bakti Sarimas", "Poktan Bakti Karya Nyata", 208, 208, "Ibul, Pucuk Rantau", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-KUANS-11", "Kuansing", "PT Surya Agrolika Reksa / Adimulia", "Agrinas", 254, 254, "Singingi Hilir", "Agrinas/KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    # PELALAWAN
    ("K-PEL-01", "Pelalawan", "Mitra Unggul Pusaka", "Kop. Karya Indragiri Maju", 366.8, 304.01, "Segati/Langgam", "Sudah KSO", "Hijau", "Nihil Intelkam", "Intelkam", "sedang"),
    ("K-PEL-02", "Pelalawan", "Sari Lembah Subur", "Maju Serempak", 681.01, 19, "Genduang", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-PEL-03", "Pelalawan", "Mekar Sari Alam Lestari", "Berlian Nusantara Perkasa", 4745, None, "Mak Teduh", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-PEL-04", "Pelalawan", "Serikat Putra", "Agrinas", 138, None, "Sialang Godang", "Agrinas", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-PEL-05", "Pelalawan", "Gandaerah Hendana", "Berlian NP", 10, 10, "Kerumutan (bagian Pelalawan)", "Sudah KSO", "Hijau", "Sebagian luas di Inhu", "Intelkam", "sedang"),
    ("K-PEL-06", "Pelalawan", "Eka Sari Lorena", "Kop. Segati Sejahtera", 127, None, "Segati", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-PEL-07", "Pelalawan", "Guna Dodos", "Maju Serempak", 5, None, "Sei Kijang", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-PEL-08", "Pelalawan", "Agrita Sari Prima", "Kop. Desa Segati", 111.82, None, "Segati", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-PEL-09", "Pelalawan", "Viktorindo Alam Lestari", "Poktan LMMB", 85, None, "Palas, PKL Kuras", "Sudah KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-PEL-10", "Pelalawan", "Agregat Agrinas Prima / TNTN", "10 KSO penunjukan Agrinas", 29000, 5392.83, "Kab. Pelalawan + TNTN ~28000 Ha masyarakat", "Campuran", "Hijau (KSO); TNTN terpisah", "Angka 180.797.180 Ha DITOLAK; TNTN aktif unjuk rasa", "Pelalawan PDF+TNTN", "tipis"),
    # BENGKALIS
    ("K-BKS-01", "Bengkalis", "PT Sinar Inti Sawit (SIS)", "PT Palma Agung Betuah (PAB)", 4557.19, 4557.19, "Bumbung / Bathin Solapan / Mandau", "Sudah KSO", "Merah", "Bentrok Des 2025–Mei 2026; pos dibakar; multi-LP", "Intelkam+Mei2026", "kuat"),
    ("K-BKS-02", "Bengkalis", "PT Surya Dumai Agrindo", "Kop. Madani Berkah Bersama", 114, 114, "Buruk Bakul, Bukit Batu", "Sudah KSO", "Hijau", "Kondusif", "Mei2026", "kuat"),
    ("K-BKS-03", "Bengkalis", "PT Priatama Riau Rupat", "Kop. Peduli Kampung Darul Aman", 83, 83, "Darul Aman, Rupat", "Sudah KSO", "Hijau", "Kondusif", "Mei2026", "kuat"),
    ("K-BKS-04", "Bengkalis", "CV Hendrik Padang", "Kop. Mitra Karya Perkasa", 749.6, 749.6, "Bandar Jaya, Siak Kecil", "Sudah KSO (belum operasi)", "Kuning*", "Penolakan KSO; panggilan Agrinas Jakarta", "Mei2026", "kuat"),
    ("K-BKS-05", "Bengkalis", "CV Sepakat Bersama Ali", "Kop. Sepahat Bersatu", 1204.9, 1204.9, "Sepahat, Bandar Laksamana", "Sudah KSO", "Kuning*", "Unjuk rasa/pengusiran; kini operasi", "Mei2026", "kuat"),
    ("K-BKS-06", "Bengkalis", "PT Sumatera Riang Lestari", "Belum ada KSO", 39003, None, "Rupat", "Belum KSO", "Hijau", "Nihil", "Intelkam", "sedang"),
    ("K-BKS-07", "Bengkalis", "PT SPM + PT BBHA", "Belum ada KSO", 1532.05, None, "Bandar Laksamana", "Belum KSO", "Hijau", "Belum tindak lanjut PKH", "Mei2026", "sedang"),
    ("K-BKS-08", "Bengkalis", "PT Mutiara Naga Indonesia", "PT Permata Kencana Utama / Agrinas", 1500, None, "Bengkalis", "Sudah KSO (sengketa)", "—", "Mediasi gagal Jul 2026; bentrok 29 Jul", "Mei2026 update", "kuat"),
    ("K-BKS-09", "Bengkalis", "PT Darmali Jaya Lestari", "Belum KSO", 671.78, None, "Air Kulim, Bathin Solapan", "Belum KSO", "—", "Belum tindak lanjut Satgas", "Mei2026", "sedang"),
    ("K-BKS-10", "Bengkalis", "Lainnya (Muriniwood, Berlian Agro, SSS, Indograha, MMJ, Murini Sam Sam)", "Campuran", None, None, "Kab. Bengkalis", "Campuran", "Hijau mayoritas", "Sebagian belum KSO", "Mei2026", "sedang"),
    # KAMPAR
    ("K-KMP-01", "Kampar", "PT Johan Sentosa", "Dikelola sendiri Agrinas", 5764, None, "Pasir Sialang, Bangkinang", "Dikelola Sendiri", "— (lokal hotspot)", "5 LP Jun 2026 pencurian/pengeroyokan terkait klaim ganda", "Kampar luas+LP", "kuat"),
    ("K-KMP-02", "Kampar", "PT Torus Ganda Tambusan Timur", "Dikelola sendiri", 12693.26, None, "Kampar (overlap Rohul naming)", "Dikelola Sendiri", "—", "Luas terbesar di daftar Kampar Agrinas", "Kampar luas", "sedang"),
    ("K-KMP-03", "Kampar", "PT PSPI", "Poktan Riau Jaya Makmur / Paruh Marta", 1445, 1445, "Petapahan & Kampar Kiri", "Sudah KSO", "Kuning", "Tumpang tindih sitaan vs masyarakat/koperasi", "Intelkam", "kuat"),
    ("K-KMP-04", "Kampar", "PT Ciliandra Perkasa", "Agrinas", 107, 107, "Siabu, Bangkinang Barat", "Sudah KSO", "Hijau", "Nihil", "Intelkam+Kampar", "kuat"),
    ("K-KMP-05", "Kampar", "PT Sarindo Agro Lestari / Kepau Jaya", "Poktan Riau Jaya Makmur", 1000, 1000, "Kepau Jaya, Siak Hulu", "Sudah KSO", "Kuning", "Penolakan Gapoktan Tiga Desa Maju", "Intelkam", "kuat"),
    ("K-KMP-06", "Kampar", "CV Jaya/Makmur Jaya Sentosa", "Poktan Kampar Jaya Bersama", 1070.59, 1070.59, "Kuali, Tambang", "Sudah KSO", "Kuning", "Tolak pernyataan Agrinas; PAM Flores non-BUJP", "Intelkam+Kampar", "kuat"),
    ("K-KMP-07", "Kampar", "PT Anugerah Tuah Mulya Perkasa", "KSO", 1891.35, None, "Kampar", "Sudah KSO", "—", "Dalam daftar KSO Kampar", "Kampar luas", "sedang"),
    ("K-KMP-08", "Kampar", "PT Riau Jaya Utama", "KSO", 1860.1, None, "Kampar", "Sudah KSO", "—", "", "Kampar luas", "sedang"),
    ("K-KMP-09", "Kampar", "PT Ramajaya Pramukti", "KSO", 1209.62, None, "Kampar", "Sudah KSO", "—", "", "Kampar luas", "sedang"),
    ("K-KMP-10", "Kampar", "PT Bina Pitri Jaya", "KSO", 895.78, None, "Kampar", "Sudah KSO", "—", "", "Kampar luas", "sedang"),
    ("K-KMP-11", "Kampar", "Lainnya Belum KSO (Arindo, Simas, Sarana Inti, Alam Riau, Hotman, Peputra, Ganda Buanindo, Inti Kamparindo)", "-", None, None, "Kampar", "Belum", "—", "8 unit Belum KSO di daftar lokal 21 estate; total Agrinas 29.174,71 Ha", "Kampar luas", "sedang"),
    # SIAK
    ("K-SIAK-01", "Siak", "PT Ivo/Ivomas Tunggal", "Tiga Bintang Sinergi", 25.44, 25.44, "Kandis", "Sudah KSO", "Hijau", "Kolom luas tabel Siak kacau", "Intelkam+Siak PDF", "tipis"),
    ("K-SIAK-02", "Siak", "PT BMI Berlian Mitra Inti", "Agung Anugerah Sawit", 960, 960, "Kandis", "Sudah KSO", "Hijau", "", "Intelkam", "sedang"),
    ("K-SIAK-03", "Siak", "PT Arara Abadi / RAPP / Tahura", "— (HTI/hutan)", None, None, "Kandis / Mandau / Minas", "Bukan KSO sawit murni", "Hijau", "Sengketa admin Setda; angka 62.193 Ha OCR suspect", "Siak PDF", "tipis"),
    ("K-SIAK-04", "Siak", "PT Sumber Sari/Seraya Lestari", "—", 117.24, None, "Minas", "—", "Hijau", "Konflik lahan + anarkis vs PT SSL 2025 (Setda)", "Siak+Intelkam", "sedang"),
    # DUMAI
    ("K-DUM-01", "Dumai", "PT Duta Mas Makmur Perkasa", "PT Riden Jaya Konstruksi", 1458.7, 1458.7, "Barak Aceh / Pelintung", "Sudah KSO", "Kuning", "Bentrok tumpang tindih lahan; take-over Feb 2026", "Intelkam+Mei2026", "kuat"),
    ("K-DUM-02", "Dumai", "PT Sinar Riau Palm Oil", "Self / KSO", 1498, 1498, "Jl Sudirman / area Dumai", "Sudah KSO", "Kuning", "Bentrok perebutan pekerjaan", "Intelkam", "sedang"),
    ("K-DUM-03", "Dumai", "PT Pelintung Jaya Bersama", "Self", 500, 500, "Mundam, Medang Kampai", "Sudah KSO", "Kuning", "Bentrok perebutan pekerjaan", "Intelkam", "sedang"),
    # MERANTI
    ("K-MER-01", "Kep. Meranti", "PT Sumatra/Sumatera Riang Lestari", "—", 2584.44, None, "Kep. Meranti (alamat OCR Pekanbaru)", "—", "Hijau", "Intelkam 1 perusahaan; laporan lokal NIHIL semua klaster", "Intelkam+Meranti DOCX", "tipis"),
    # PEKANBARU
    ("K-PKU-01", "Pekanbaru", "5 perusahaan sawit (Satria Windu, Budi Tani, Sawit Unggul, Surya Intisari, Aneka Inti)", "NIHIL KSO/Agrinas", None, None, "Wilkum Pekanbaru", "NIHIL Agrinas/KSO", "NIHIL", "Tidak ada sitaan PKH/Agrinas/KSO tercatat", "Bismillah", "sedang"),
]

for r in kebun_rows:
    ws.append(list(r))
style_header(ws, len(kebun_headers))
autosize(ws)

# ===================== SHEET: kasus_lp =====================
ws2 = wb.create_sheet("kasus_lp")
kasus_headers = [
    "id", "polres", "nomor_lp", "tanggal", "kategori", "pihak", "lokasi",
    "uraian_singkat", "dampak", "status_proses", "kait_agrinas_kso", "sumber", "kepercayaan",
]
ws2.append(kasus_headers)

kasus_rows = [
    ("C-ROHUL-01", "Rohul", "LP/B/07/II/2026/SPKT/POLSEK BONAI DARUSSALAM/...", "2026-02-07", "Pembunuhan/penyerangan PAM", "PAM Nusantara Sawit Majuma vs PAM KUD Telago Biru / Berkat Satu", "Sontang, Bonai Darussalam", "Serangan massal barak/mess terkait klaim KSO Agrinas", "1 MD + 6 luka", "P21 / multi-tersangka", "Ya - KSO Majuma", "Rohul+Intelkam", "kuat"),
    ("C-ROHUL-02", "Rohul", "LP/B/107/III/2026/SPKT/POLRES ROKAN HULU/...", "2026-03-31", "Pencurian TBS", "D. Reguel Parhusip vs M. Jamil Siregar / RKG", "Agrinas eks Torganda Afd XII, Tambusai Utara", "Hentikan panen Agrinas; ambil ~208 tandan", "Kerugian TBS ~2 ton", "P21", "Ya", "Rohul DOCX", "kuat"),
    ("C-ROHUL-03", "Rohul", "LP/B/32/I/2026/SPKT/POLRES ROKAN HULU/...", "2026-01-21", "Penganiayaan", "Ali Dahlan Lubis vs Zulherman Syarif", "Afd IX PT Agrinas Palma Nusantara", "Hadang patroli Agrinas+Satgas PKH", "Luka", "P21", "Ya", "Rohul DOCX", "kuat"),
    ("C-ROHUL-04", "Rohul", "LP/B/21/II/2026/SPKT/POLSEK TAMBUSAI/...", "2026-02-22", "Pengrusakan / paksa masuk", "Leonardo Marbun vs Hose Fernando cs + massa ~400", "Mess Afd VIII Kebun Tambusai Timur", "Massa buka paksa kantor/mess; rusak kendaraan", "Rugi Rp5jt (kendaraan)", "P21", "Ya - Agrinas/Torganda", "Rohul DOCX", "kuat"),
    ("C-ROHIL-01", "Rohil", "RJ (bentrok 20 Okt 2025)", "2025-10-20", "Bentrok panen sepihak", "W. Siringo-Ringo vs security Ujung Tanjung Sejahtera", "Blok 29-30 Ex Rumbia I, Gunung Mas Raya", "Panen sepihak TBS vs pembubaran security", "7 luka", "RJ selesai", "Ya - KSO UTS", "Intelkam", "kuat"),
    ("C-BKS-01", "Bengkalis", "476/XII/2025/...SEK MDU + LP/B/153/XII/2025", "2025-12-22", "Penganiayaan/pengrusakan", "PT PAB vs karyawan PT SIS", "Lahan sitaan SIS Mandau", "Bentrok perebutan lahan sitaan PKH", "3 luka; 13 mobil dirusak", "Lidik/sidik", "Ya - KSO PAB", "Mei2026", "kuat"),
    ("C-BKS-02", "Bengkalis", "Kejadian 3 Des 2025", "2025-12-03", "Bentrok", "Karyawan PT SIS vs masyarakat Sakai", "Bengkalis/Mandau", "Pertahankan TBS", "1 warga Sakai luka kepala", "—", "Ya - areal SIS", "Intelkam", "sedang"),
    ("C-BKS-03", "Bengkalis", "Kejadian 14 Jan 2026", "2026-01-14", "Pembakaran pos", "Warga Bukit Abas vs security PAB", "Pos 1 PAB", "Cekcok knalpot → bakar pos", "2 LP; 5 tersangka", "Penyelidikan", "Ya - KSO PAB", "Intelkam", "kuat"),
    ("C-BKS-04", "Bengkalis", "LP/219/V/2026 + LP/B/59/V/2026", "2026-05-16", "Pencurian TBS + penganiayaan/pembakaran", "PAB vs kubu Rusman Tobing / Yakub Saragih cs", "Lahan sitaan PAB Mandau", "Panen liar ~150 Ha; 6 motor dibakar", "Ketegangan massa", "2 diamankan", "Ya", "Mei2026", "kuat"),
    ("C-BKS-05", "Bengkalis", "Kejadian 28-30 Jul 2026", "2026-07-29", "Bentrok massa KSO", "Agrinas+vendor PKU vs kelompok tani Mutiara Naga", "Mutiara Naga Indonesia", "Mediasi gagal; bentrok masuk lokasi", "Korban luka kedua belah pihak", "Intelkam monitoring", "Ya", "Mei2026 update", "kuat"),
    ("C-DUM-01", "Dumai", "— (narasi Intelkam)", "2026-02-20", "Bentrok / take-over", "DMMP vs Riden Jaya Konstruksi", "Pelintung, Medang Kampai", "Tumpang tindih lahan / take-over KSO", "Bentrok fisik", "—", "Ya", "Intelkam+Mei2026", "sedang"),
    ("C-KMP-01", "Kampar", "LP/B/165/VI/2026/...POLRES KAMPAR", "2026-06-06", "Pencurian dari sengketa lahan", "Nocolas Hutabarat vs Aroel Campay dkk", "Divisi V PT Johan Sentosa, Pasir Sialang", "Pencurian buah sawit terkait klaim ganda", "—", "Lidik", "Ya - Agrinas Johan Sentosa", "Kampar Agrinas 2026", "kuat"),
    ("C-KMP-02", "Kampar", "LP/B/168/VI/2026/...", "2026-06-07", "Pengeroyokan", "Tolo Zatulo Zega vs Yudi dkk", "Pos Gudang Kantor Agrinas Johan Sentosa", "Korban tangkap pencuri lalu dikeroyok", "Luka", "Sidik", "Ya", "Kampar", "kuat"),
    ("C-KMP-03", "Kampar", "LP/B/169/VI/2026/...", "2026-06-07", "Pencurian", "Gatili Lase vs Fais dkk", "PT Agrinas Sei Jernih, Pasir Sialang", "Pencurian di kebun Agrinas", "—", "Sidik", "Ya", "Kampar", "kuat"),
    ("C-KMP-04", "Kampar", "LP/B/190/VI/2026/...", "2026-06-21", "Pencurian", "Gatili Lase vs Dogol Siregar", "Kebun PT Agrinas Palma Nusantara, Pasir Sialang", "Pencurian TBS", "—", "Sidik", "Ya", "Kampar", "kuat"),
    ("C-KMP-05", "Kampar", "LP/B/197/VI/2026/...", "2026-06-28", "Pencurian", "Gatili Lase vs Dandi Saputra", "Kebun PT Agrinas Palma Nusantara, Pasir Sialang", "Pencurian TBS", "—", "Sidik", "Ya", "Kampar", "kuat"),
    ("C-KMP-06", "Kampar", "Agregat potensi konflik agraria 2024-2026", "2024-2026", "Penyerobotan/pengrusakan/pengaduan", "Beragam (masyarakat antar pihak + korporasi)", "Kab. Kampar", "22 (2026) + 27 (2025) + 20 (2024) entri LP/pengaduan; mayoritas bukan langsung Agrinas", "—", "Campuran", "Sebagian", "Kampar konflik DOCX", "kuat"),
    ("C-INHIL-01", "Inhil", "Laporan 20 Mei 2026 (belum nomor di teks)", "2026-05-20", "Pemanenan/pencurian TBS", "PT RSA vs duga KSO Cipta Nugraha", "Blok M65 Div 2 Batang Sari, Mandah", "~1,1 ton TBS dipanen non-karyawan", "Penyelidikan", "Penyelidikan", "Ya", "Inhil PDF", "kuat"),
    ("C-INHU-01", "Inhu", "LP pencurian TBS (nomor tidak di tabel)", "2026-04", "Sengketa + panen paksa + LP", "KSO TKBM vs Asli Hutagaol / Adv Polman Sinaga", "EX Palm Lestari Makmur, Penyaguhan", "Sengketa 560 Ha; bentrok panen", "Bentrok", "—", "Ya", "INHU DOCX", "sedang"),
    ("C-INHU-02", "Inhu", "Aksi/mediasi 8 Jun 2026", "2026-06-08", "Aksi tuntutan kelola bersama", "PKN Kasmiran / Poktan JD Karya Mandiri vs TPP", "Redang Seko / Lirik", "Minta kelola bersama lahan KSO", "Mediasi", "Mediasi", "Ya", "INHU DOCX", "sedang"),
    ("C-KUANS-01", "Kuansing", "~35-37 LP Agrinas-relevant 2020-2026", "2020-2026", "Mayoritas pencurian TBS (+1 penganiayaan)", "Pelaku individu vs kebun Cerenti Subur / WJT / PTPN IV / Agrinas", "Cerenti, Pesikaian, Inuman, Sigaruntang", "Pencurian berulang di areal yang kemudian dikelola Agrinas", "Kerugian kecil-sedang; beberapa angka 2024 DITOLAK", "Campuran tipiring/selesai/lidik", "Ya (areal)", "Kuansing LP PDF", "kuat"),
    ("C-KUANS-02", "Kuansing", "Potensi: Pesikaian TNTN/PTPN", "2026", "Potensi konflik relokasi", "Warga Pesikaian vs Koptan Bagan Limau vs Satgas/PTPN/Agrinas", "Pesikaian, Cerenti", "Bagi hasil & relokasi lahan sitaan tidak clear", "Klaster kuning", "Proses penyelesaian", "Ya", "Potensi 2026", "kuat"),
    ("C-KUANS-03", "Kuansing", "Potensi: WJT 20%", "2026", "Tuntutan pemanfaatan", "Masyarakat Cerenti & Inuman vs Agrinas/vendor", "WJT area", "Tuntut 20% via koperasi; komunikasi tertutup", "Potensi", "Proses Satgas", "Ya", "Potensi 2026", "kuat"),
    ("C-PEL-01", "Pelalawan", "LP KSO vs pihak lain (nomor tidak disebut)", "—", "Pencurian / halangi panen", "KSO vs pihak lain; masyarakat vs KSO Bunut", "Kec. Bunut dkk", "Saling lapor; di Bunut sudah berdamai", "Berdamai (Bunut)", "Damai parsial", "Ya", "Pelalawan PDF", "tipis"),
    ("C-PEL-02", "Pelalawan", "Register aksi TNTN (bukan LP tunggal)", "2025-2026", "Unjuk rasa / penolakan portal / pengusiran Satgas", "AMMP, FKPM, Forum Tata Kelola, LMND, KOMARI, ACER vs penertiban TNTN", "7 desa TNTN Pelalawan", "8 unjuk rasa; 3 tolak portal; 5 usir Satgas; 2 spontan; 1 intimidasi", "6 tersangka perusakan; 3 tersangka KSDAE", "Relatif kondusif pasca penangkapan (Feb 2026)", "Terkait kawasan; bukan KSO murni", "TNTN 23.02", "kuat"),
    ("C-SIAK-01", "Siak", "Rekap Setda 2024-2025 (10 entri admin)", "2024-2025", "Sengketa administratif/batas/izin", "Masyarakat vs Arara/RAPP/SSL/PTPN V/koperasi", "Berbagai kampung Siak", "Mayoritas bukan KSO Agrinas sawit; satu kasus anarkis bakar kantor PT SSL", "Pembakaran kantor SSL", "Mediasi/RDP", "Lemah kait Agrinas", "Siak PDF", "sedang"),
    ("C-ROHIL-02", "Rohil", "Agregat LP Agrinas=3; masyarakat LP=20 / pengaduan=194", "—", "Campuran", "Agrinas vs masyarakat vs perusahaan", "Kab. Rokan Hilir", "Detail nomor LP tidak tersedia di sumber", "—", "—", "Ya (agregat)", "Bismillah", "tipis"),
    ("C-PKU-01", "Pekanbaru", "7 LP pencurian sawit Polsek Rumbai/Rumbai Barat 2025-2026", "2025-2026", "Pencurian TBS masyarakat", "Pelaku individu vs kebun pribadi/perusahaan non-Agrinas", "Rumbai / Rumbai Barat", "Pidana oportunistik; Intelkam sebut TP Agrinas NIHIL", "Rugi kecil; RJ/tipiring", "Campuran", "Tidak (bukan Agrinas/KSO)", "Pencurian DOCX", "kuat"),
    ("C-MER-01", "Kep. Meranti", "NIHIL", "—", "NIHIL", "—", "—", "Semua klaster konflik NIHIL di laporan lokal", "—", "—", "NIHIL", "Meranti DOCX", "sedang"),
]

for r in kasus_rows:
    ws2.append(list(r))
style_header(ws2, len(kasus_headers))
autosize(ws2)

# ===================== SHEET: ringkasan_polres =====================
ws3 = wb.create_sheet("ringkasan_polres")
ring_headers = [
    "polres", "jml_kebun_intelkam", "merah", "kuning", "hijau", "konflik_resume",
    "sinyal_konflik", "kualitas_data", "tipologi_dominan", "prioritas_harda", "catatan",
]
ws3.append(ring_headers)
ring_rows = [
    ("Inhu", 28, 0, 2, 26, 3, "Tinggi (volume+narasi)", "Kuat", "Plasma/MHA/tuntutan kelola; bentrok pemanen", "Tinggi", "Volume kebun tertinggi; LP formal jarang di sumber lokal"),
    ("Rohil", 24, 1, 6, 17, 7, "Sangat tinggi", "Sedang", "Penolakan KSO; bentrok panen; PAM non-BUJP", "Sangat tinggi", "1 merah (UTS); detail LP Agrinas tipis di Bismillah"),
    ("Rohul", 21, 1, 5, 15, 6, "Sangat tinggi (MD)", "Kuat", "Bentrok PAM swakarsa; penolakan KSO; panen sepihak Torganda", "Sangat tinggi", "Kasus MD Majuma; 12 KSO; 4 LP P21"),
    ("Inhil", 15, 0, 4, 11, 4, "Sedang", "Sedang", "Tuntutan kembalikan kelola; pemanenan KSO vs eks-perusahaan", "Sedang-tinggi", "Intelkam lebih kaya dari PDF konflik lokal (hanya 2 baris)"),
    ("Kuansing", 11, 0, 1, 10, 1, "Tinggi (pidana berulang)", "Kuat", "Pencurian TBS berulang + potensi TNTN/Pesikaian + tuntutan 20% WJT", "Tinggi", "Log LP 2020-2026 kaya; bedakan pidana vs struktural"),
    ("Pelalawan", 9, 0, 0, 9, 0, "Sedang (TNTN tinggi)", "Tipis (KSO) / Kuat (TNTN)", "TNTN kawasan hutan; LP KSO tipis; semua KSO hijau", "Tinggi (dimensi TNTN)", "Pisahkan analisis KSO hijau vs konflik TNTN"),
    ("Bengkalis", 8, 1, 0, 7, 3, "Sangat tinggi", "Kuat", "Bentrok KSO PAB-SIS; penolakan KSO; Mutiara Naga Jul 2026", "Sangat tinggi", "Inkonsistensi klaster kuning pada kartu vs resume"),
    ("Siak", 6, 0, 0, 6, 1, "Rendah-sedang (admin)", "Tipis", "Sengketa administratif/HTI lebih dominan ketimbang KSO Agrinas", "Sedang", "Tabel luas tidak dapat dijumlah"),
    ("Kampar", 4, 0, 3, 1, 3, "Tinggi (LP bersih)", "Kuat", "Klaim ganda + pencurian di Johan Sentosa; penolakan Gapoktan/KSO", "Tinggi", "Intelkam 4 kebun vs lokal 21 estate Agrinas 29rb Ha"),
    ("Dumai", 3, 0, 3, 0, 3, "Tinggi", "Sedang", "Bentrok perebutan kerja/lahan antar pengelola", "Tinggi", "DMMP overlap pencatatan dengan Bengkalis"),
    ("Kep. Meranti", 1, 0, 0, 1, 0, "Nihil/rendah", "Tipis", "NIHIL lokal; 1 perusahaan Intelkam hijau", "Rendah (gap-fill)", "Jangan over-infer dari NIHIL"),
    ("Pekanbaru", 0, 0, 0, 0, 0, "Nihil Agrinas", "Sedang", "Pencurian TBS masyarakat non-Agrinas", "Rendah (monitoring)", "Bedakan pidana biasa vs konflik Agrinas-KSO"),
]
for r in ring_rows:
    ws3.append(list(r))
style_header(ws3, len(ring_headers))
autosize(ws3)

# ===================== SHEET: data_quality =====================
ws4 = wb.create_sheet("data_quality")
dq_headers = ["id", "isu", "polres_terkait", "dampak_analisis", "tindakan"]
ws4.append(dq_headers)
dq_rows = [
    ("DQ01", "Duplikat file DATA LAHAN... MEI 2026 (1).pdf", "Bengkalis/Dumai", "Noise inventaris", "Pakai satu file saja"),
    ("DQ02", "Filename POTENSI KONFLIK AGRARIA POLDA RIAU isi hanya Kampar", "Kampar", "Salah persepsi cakupan", "Label ulang sebagai Kampar-only"),
    ("DQ03", "Rohil total 20 vs 24; 1+6+17=24", "Rohil", "Inkonsistensi hitungan", "Pakai 24 dari breakdown klaster"),
    ("DQ04", "Kampar resume H3/K1 vs kartu K3/H1", "Kampar", "Salah klaster", "Pakai kartu: 3 kuning 1 hijau"),
    ("DQ05", "Bengkalis resume 0 kuning vs kartu 3 kuning", "Bengkalis", "Underestimate risiko", "Flag kuning* dari kartu/Mei2026"),
    ("DQ06", "Pelalawan luas 180.797.180 ha implausible", "Pelalawan", "Jangan pakai angka", "Tolak; minta klarifikasi Disbun"),
    ("DQ07", "Siak kolom luas keseluruhan/KSO/Agrinas kacau", "Siak", "Tidak bisa agregasi Ha", "Pakai Intelkam cards saja"),
    ("DQ08", "Kuansing kerugian 2024 tipiring ratusan juta", "Kuansing", "Overstate kerugian", "Tolak field kerugian 2024 mencurigakan"),
    ("DQ09", "Arara Abadi sita 62.193 ha OCR suspect", "Siak", "Overstate", "Verifikasi BPKH/Disbun"),
    ("DQ10", "DMMP tercatat Bengkalis dan Dumai", "Bengkalis; Dumai", "Double count risiko", "Satu record, flag multi-wilayah"),
    ("DQ11", "Kampar Table2 jumlah 636.366 ≠ masyarakat+Agrinas", "Kampar", "Total kab invalid", "Jangan pakai total untuk rasio"),
    ("DQ12", "Tebing Tinggi DOCX corrupt CRC", "Kep. Meranti", "Tidak terbaca", "Minta kirim ulang"),
    ("DQ13", "Majuma tersangka count beda 12.02 vs 23.02", "Rohul", "Inkonsistensi hukum", "Pakai update terbaru + cek LP"),
    ("DQ14", "INHIL narasi sebut 2 merah vs angka 0 merah", "Inhil", "Salah label", "Pakai angka 0M/4K/11H"),
    ("DQ15", "Banyak hijau Intelkam punya konflik di sumber lokal", "Inhu; Kuansing", "Klaster understate", "Validasi ulang klaster vs kejadian 2025-2026"),
]
for r in dq_rows:
    ws4.append(list(r))
style_header(ws4, len(dq_headers))
autosize(ws4)

# ===================== SHEET: agregat_polda =====================
ws5 = wb.create_sheet("agregat_polda")
ws5.append(["metrik", "nilai", "sumber", "catatan"])
agg = [
    ("Total perusahaan KSO/PKH Riau", 130, "Intelkam 12.02.2026", "Resume resmi"),
    ("Klaster Merah", 3, "Intelkam", "2% - Rohil, Rohul, Bengkalis"),
    ("Klaster Kuning", 24, "Intelkam", "18%"),
    ("Klaster Hijau", 103, "Intelkam", "80% - perlu validasi vs kejadian aktual"),
    ("Bentrok menonjol", 6, "Intelkam", "Rohil, Rohul, Bengkalis"),
    ("Korban jiwa/luka (resume)", "1 MD + 23 luka", "Intelkam", "Total 24 orang"),
    ("Perusahaan tanpa BUJP (named)", 6, "Intelkam", "PAM swakarsa bermasalah"),
    ("Satker dengan kebun Agrinas/KSO", "11 dari 12", "Intelkam", "Pekanbaru NIHIL dalam daftar 130"),
    ("Periode data dominan", "2024 - Jul 2026", "Multi-sumber", "LP Kuansing back to 2020"),
]
for r in agg:
    ws5.append(list(r))
style_header(ws5, 4)
autosize(ws5)

out = r"C:\Users\Patron\Downloads\sawit lagi\matriks_agrinas_kso_12_polres.xlsx"
wb.save(out)
print(f"Saved {out}")
print(f"kebun={len(kebun_rows)} kasus={len(kasus_rows)} ringkasan={len(ring_rows)} dq={len(dq_rows)}")
