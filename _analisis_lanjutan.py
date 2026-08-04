# -*- coding: utf-8 -*-
"""5 analisis lanjutan Unit II Harda — skor risiko, validasi klaster,
jaringan aktor, timeline eskalasi, gap-fill."""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import re

BASE = r"C:\Users\Patron\Downloads\sawit lagi"
SRC = f"{BASE}\\matriks_agrinas_kso_12_polres.xlsx"
OUT_XLSX = f"{BASE}\\analisis_lanjutan_5_prioritas.xlsx"
OUT_MD = f"{BASE}\\laporan_5_analisis_lanjutan.md"

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(color="FFFFFF", bold=True, size=10)
red_fill = PatternFill("solid", fgColor="FF6B6B")
orange_fill = PatternFill("solid", fgColor="FFD93D")
yellow_fill = PatternFill("solid", fgColor="FFF3B0")
green_fill = PatternFill("solid", fgColor="C6EFCE")
thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(1, col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = thin


def autosize(ws, max_width=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value:
                length = max(length, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(11, length + 2)


def load_sheet_dicts(path, sheet):
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h else f"col{i}" for i, h in enumerate(rows[0])]
    out = []
    for r in rows[1:]:
        if not any(r):
            continue
        out.append(dict(zip(headers, r)))
    return out


kebun = load_sheet_dicts(SRC, "kebun")
kasus = load_sheet_dicts(SRC, "kasus_lp")
ring = load_sheet_dicts(SRC, "ringkasan_polres")

# ===================== R1 SCORING =====================
# Weights (documented): klaster 30, sinyal konflik 25, korban/MD 20,
# PAM non-BUJP 15, kualitas data penalty inverted as completeness 10

KLAS_SCORE = {"Merah": 30, "Kuning": 18, "Hijau": 4, "—": 8, "Campuran": 12}
SINYAL_SCORE = {
    "Sangat tinggi (MD)": 25,
    "Sangat tinggi": 22,
    "Tinggi (volume+narasi)": 18,
    "Tinggi (pidana berulang)": 16,
    "Tinggi (LP bersih)": 16,
    "Tinggi": 15,
    "Sedang (TNTN tinggi)": 14,
    "Sedang": 10,
    "Rendah-sedang (admin)": 6,
    "Nihil/rendah": 2,
    "Nihil Agrinas": 1,
}
PRIO_BONUS = {
    "Sangat tinggi": 0,  # already in signals
    "Tinggi": 0,
    "Tinggi (dimensi TNTN)": 0,
    "Sedang-tinggi": 0,
    "Sedang": 0,
    "Rendah (gap-fill)": 0,
    "Rendah (monitoring)": 0,
}

# Known PAM non-BUJP from Intelkam
PAM_NON_BUJP = {
    "Rohul": ["Nusantara Sawit Majuma", "Torus", "Togos", "Togas"],
    "Bengkalis": ["Palma Agung Betuah", "PAB"],
    "Rohil": ["Ujung Tanjung Sejahtera", "UTS"],
    "Kampar": ["Makmur Jaya Sentosa", "Jaya Makmur"],
}

KORBAN_POLRES = {
    "Rohul": 25,   # 1 MD weighted heavily + luka
    "Rohil": 12,   # 7 luka
    "Bengkalis": 14,  # multi-event luka + kerusakan
    "Dumai": 6,
    "Kampar": 4,
    "Inhu": 5,
    "Kuansing": 3,
    "Inhil": 3,
    "Pelalawan": 8,  # TNTN aksi though not KSO deaths
    "Siak": 4,  # SSL burning
    "Kep. Meranti": 0,
    "Pekanbaru": 0,
}


def klaster_score_polres(row):
    m = int(row.get("merah") or 0)
    k = int(row.get("kuning") or 0)
    h = int(row.get("hijau") or 0)
    total = m + k + h
    if total == 0:
        return 0
    # weighted share * 30 max
    return round((m * 30 + k * 18 + h * 4) / total, 1)


def completeness_score(kualitas):
    q = (kualitas or "").lower()
    if "kuat" in q:
        return 10
    if "sedang" in q:
        return 6
    if "tipis" in q:
        return 3
    if "nihil" in q:
        return 2
    return 5


satker_scores = []
for r in ring:
    pol = r["polres"]
    jml = int(r.get("jml_kebun_intelkam") or 0)
    s_klas = klaster_score_polres(r)
    s_sinyal = SINYAL_SCORE.get(r.get("sinyal_konflik"), 8)
    s_korban = min(20, KORBAN_POLRES.get(pol, 0))
    # PAM: 15 if known non-BUJP in polres
    s_pam = 15 if pol in PAM_NON_BUJP else (8 if jml >= 10 else 2)
    s_data = completeness_score(r.get("kualitas_data"))
    # volume factor small boost
    s_vol = min(8, jml * 0.25)
    total = round(s_klas + s_sinyal + s_korban + s_pam + s_data * 0.5 + s_vol, 1)
    # normalize-ish cap display 100
    total_100 = min(100, round(total / 1.08, 1))
    if total_100 >= 70:
        band = "KRITIS"
    elif total_100 >= 55:
        band = "TINGGI"
    elif total_100 >= 35:
        band = "SEDANG"
    else:
        band = "RENDAH"
    satker_scores.append({
        "polres": pol,
        "jml_kebun": jml,
        "skor_klaster": s_klas,
        "skor_sinyal": s_sinyal,
        "skor_korban": s_korban,
        "skor_pam_non_bujp": s_pam,
        "skor_kelengkapan_data": round(s_data * 0.5, 1),
        "skor_volume": round(s_vol, 1),
        "skor_total": total_100,
        "band_risiko": band,
        "tipologi_dominan": r.get("tipologi_dominan"),
        "catatan": r.get("catatan"),
    })

satker_scores.sort(key=lambda x: -x["skor_total"])

# Kebun-level scoring
kebun_scores = []
for k in kebun:
    pol = k.get("polres")
    klas = str(k.get("klaster") or "—")
    if "Merah" in klas:
        sk = 40
    elif "Kuning" in klas:
        sk = 28
    elif "Hijau*" in klas or "Hijau" in klas:
        # boost if konflik note suggests active
        note = str(k.get("konflik_note") or "").lower()
        if any(w in note for w in ["bentrok", "md", "luka", "penolakan", "tuntutan", "panen paksa", "hotspot", "lp "]):
            sk = 18  # hijau bergesekan
        elif "nihil" in note:
            sk = 5
        else:
            sk = 8
    else:
        sk = 12

    note = str(k.get("konflik_note") or "")
    note_l = note.lower()
    if "1 md" in note_l or " md" in note_l:
        sk += 25
    elif "luka" in note_l or "bentrok" in note_l:
        sk += 15
    elif "penolakan" in note_l or "tuntutan" in note_l:
        sk += 10
    elif "lp" in note_l or "pencurian" in note_l:
        sk += 6

    # PAM non-BUJP mention
    penerima = str(k.get("penerima_kso") or "")
    eks = str(k.get("eks_perusahaan") or "")
    blob = (penerima + " " + eks + " " + note).lower()
    pam_hit = any(
        any(x.lower() in blob for x in names)
        for names in PAM_NON_BUJP.values()
    )
    if pam_hit or "non-bujp" in note_l or "pam" in note_l:
        sk += 12

    if str(k.get("status_kso") or "").lower().startswith("belum"):
        sk += 5

    sk = min(100, sk)
    if sk >= 70:
        band = "KRITIS"
    elif sk >= 50:
        band = "TINGGI"
    elif sk >= 30:
        band = "SEDANG"
    else:
        band = "RENDAH"

    kebun_scores.append({
        "id": k.get("id"),
        "polres": pol,
        "eks_perusahaan": k.get("eks_perusahaan"),
        "penerima_kso": k.get("penerima_kso"),
        "klaster_intelkam": klas,
        "status_kso": k.get("status_kso"),
        "skor_kebun": sk,
        "band_risiko": band,
        "konflik_note": note[:200],
        "kepercayaan": k.get("kepercayaan"),
    })

kebun_scores.sort(key=lambda x: -x["skor_kebun"])

# ===================== R3 CLUSTER VALIDATION =====================
# Manual curated: Intelkam cluster vs observed 2025-2026 evidence
validasi = [
    # id, polres, estate, klaster_resmi, bukti_aktual, klaster_usulan, status, alasan
    ("V01", "Rohul", "PT BK1 / Berkat Satu – Majuma", "Merah", "Bentrok 12 Jan & 7 Feb 2026; 1 MD + 6 luka; LP/B/07/II/2026", "Merah", "VALID", "Selaras dengan definisi merah"),
    ("V02", "Rohil", "PT Gunung Mas Raya – UTS", "Merah", "Bentrok 20 Okt 2025; 7 luka; RJ", "Merah", "VALID", "Selaras"),
    ("V03", "Bengkalis", "PT SIS – PAB", "Merah", "Bentrok Des 2025–Mei 2026; multi-LP; pos dibakar; kerusakan mobil", "Merah", "VALID", "Selaras; eskalasi berulang"),
    ("V04", "Bengkalis", "CV Hendrik Padang – Mitra Karya", "Hijau (resume) / Kuning* (kartu)", "Penolakan KSO; belum operasi; panggilan Agrinas Jakarta", "Kuning", "KOREKSI_NAIK", "Resume understate; kartu/Mei2026 tunjukkan konflik tanpa kerusuhan massal berulang"),
    ("V05", "Bengkalis", "CV Sepakat Bersama Ali", "Hijau (resume) / Kuning* (kartu)", "Unjuk rasa; pengusiran security; kini operasi", "Kuning", "KOREKSI_NAIK", "Sama seperti V04"),
    ("V06", "Bengkalis", "PT Mutiara Naga – PKU/Agrinas", "Tidak di kartu merah/kuning resmi", "Mediasi gagal 28 Jul; bentrok 29 Jul 2026; korban luka", "Kuning→pantau Merah", "KANDIDAT_NAIK", "Kejadian pasca briefing 12.02; perlu update klaster"),
    ("V07", "Dumai", "PT DMMP – Riden Jaya", "Kuning", "Penghadangan; take-over 20 Feb 2026; bentrok", "Kuning", "VALID", "Selaras; overlap pencatatan BKS–Dumai"),
    ("V08", "Dumai", "PT Sinar Riau Palm Oil", "Kuning", "Bentrok perebutan pekerjaan", "Kuning", "VALID", "Selaras"),
    ("V09", "Dumai", "PT Pelintung Jaya Bersama", "Kuning", "Bentrok perebutan pekerjaan", "Kuning", "VALID", "Selaras"),
    ("V10", "Kampar", "PT PSPI", "Kuning", "Tumpang tindih sitaan vs masyarakat/koperasi", "Kuning", "VALID", "Selaras"),
    ("V11", "Kampar", "PT Sarindo / Kepau Jaya", "Kuning", "Penolakan Gapoktan", "Kuning", "VALID", "Selaras"),
    ("V12", "Kampar", "CV Makmur Jaya Sentosa", "Kuning", "Tolak pernyataan Agrinas; PAM Flores non-BUJP", "Kuning", "VALID", "Selaras"),
    ("V13", "Kampar", "PT Johan Sentosa (lokal hotspot)", "Tidak masuk 4 kartu Intelkam sebagai kuning/merah", "5 LP Jun 2026 pencurian/pengeroyokan klaim ganda di areal Agrinas", "Kuning", "KANDIDAT_NAIK", "Diskrepansi Intelkam 4 kebun vs lokal 21 estate"),
    ("V14", "Inhu", "PT Indrawan Perkasa", "Kuning", "Bentrok pemanen; tutup jalan TBS", "Kuning", "VALID", "Selaras"),
    ("V15", "Inhu", "PT Selantai Argo Lestari", "Kuning", "Konflik MHA Talang Mamak", "Kuning", "VALID", "Selaras"),
    ("V16", "Inhu", "EX Palm Lestari – Koperasi TKBM", "Hijau*", "Sengketa 560 Ha; panen paksa Apr 2026; bentrok; LP pencurian", "Kuning", "KOREKSI_NAIK", "Hijau understate; ada bentrok aktual"),
    ("V17", "Inhu", "PT Tunggal Perkasa – JD Karya Mandiri", "Hijau", "Aksi/mediasi Jun 2026 minta kelola bersama", "Hijau pantau / Kuning ringan", "KANDIDAT_NAIK", "Belum kerusuhan; tuntutan aktif"),
    ("V18", "Kuansing", "PTPN IV Pesikaian", "Kuning", "Relokasi TNTN vs warga; potensi 2026 aktif", "Kuning", "VALID", "Selaras"),
    ("V19", "Kuansing", "PT Wana Jingga Timur", "Hijau", "Tuntutan 20%; komunikasi tertutup; banyak LP pencurian", "Kuning (struktural) / Hijau-pidana", "KANDIDAT_NAIK", "Pisahkan pidana TBS vs tuntutan kelola; struktural → kuning"),
    ("V20", "Kuansing", "PT Cerenti Subur", "Hijau", "Banyak LP pencurian TBS; belum KSO kelola Agrinas", "Hijau (pidana) pantau", "VALID_DENGAN_CATATAN", "Pidana berulang belum = konflik struktural; pantau jika muncul penolakan KSO"),
    ("V21", "Kuansing", "PT Merauke Tetap Jaya", "Kuning", "Status lahan / isu pencurian TBS", "Kuning", "VALID", "Selaras"),
    ("V22", "Rohil", "PT Salim Ivomas – Digjaya", "Kuning", "Penolakan penyerahan KSO", "Kuning", "VALID", "Selaras"),
    ("V23", "Rohil", "PT APSL – Satahi / KTMT", "Kuning", "Penolakan vs Agrinas/KSO", "Kuning", "VALID", "Selaras"),
    ("V24", "Rohul", "PT Torganda Tambusai Timur", "Kuning", "Pengusiran panen; LP pengrusakan mess Feb 2026", "Kuning", "VALID", "Selaras; dekat eskalasi"),
    ("V25", "Rohul", "PT Ekaudra – CV Ginting", "Kuning", "Penolakan KSO 5 desa", "Kuning", "VALID", "Selaras"),
    ("V26", "Inhil", "PT Agro Sarimas – Citra Mutiara", "Kuning", "Tuntutan kembalikan ke KPCH", "Kuning", "VALID", "Selaras"),
    ("V27", "Inhil", "PT RSA – Cipta Nugraha", "Tidak berkaster jelas di resume", "Pemanenan TBS Mei 2026; penyelidikan", "Kuning ringan", "KANDIDAT_NAIK", "Perlu masuk radar klaster satker"),
    ("V28", "Pelalawan", "9 KSO Intelkam", "Hijau semua", "LP KSO tipis; Bunut berdamai", "Hijau (KSO)", "VALID_DENGAN_CATATAN", "Portofolio TNTN terpisah — jangan campur skor KSO"),
    ("V29", "Pelalawan", "TNTN (bukan kartu KSO)", "N/A", "8 unjuk rasa; pengusiran Satgas; tersangka", "Portofolio terpisah: TINGGI", "PISAHKAN", "Jangan pakai klaster KSO untuk TNTN"),
    ("V30", "Siak", "6 kebun hijau", "Hijau", "Konflik admin/HTI Setda; bukan KSO sawit murni", "Hijau (KSO) / admin terpisah", "VALID_DENGAN_CATATAN", "Filter HTI keluar dari skor Agrinas"),
    ("V31", "Kep. Meranti", "1 perusahaan hijau", "Hijau", "Laporan lokal NIHIL", "Hijau / data gap", "TIDAK_DAPAT_DIVALIDASI", "NIHIL lokal ≠ bukti aman mutlak"),
    ("V32", "Pekanbaru", "NIHIL Agrinas", "N/A", "7 LP pencurian non-Agrinas", "N/A", "DI LUAR OBJEK", "Jangan masuk validasi klaster Agrinas"),
]

status_count = defaultdict(int)
for v in validasi:
    status_count[v[6]] += 1

# ===================== R2 ACTOR NETWORK =====================
# Normalize actor names from kebun penerima_kso + kasus
def split_actors(text):
    if not text:
        return []
    text = str(text)
    parts = re.split(r"[/;,+]|\bv[sS]\b|\bvs\b", text)
    cleaned = []
    for p in parts:
        p = p.strip()
        p = re.sub(r"\s+", " ", p)
        if len(p) < 3:
            continue
        if p.lower() in {"-", "—", "nihil", "self", "agrinas", "pt agrinas"}:
            continue
        cleaned.append(p)
    return cleaned


# Canonical hubs (seed + discovered)
actor_edges = []  # (aktor, polres, estate, peran)
actor_polres = defaultdict(set)
actor_estates = defaultdict(set)

for k in kebun:
    pol = k.get("polres")
    estate = str(k.get("eks_perusahaan") or "")[:80]
    for a in split_actors(k.get("penerima_kso")):
        # skip generic
        if a.lower() in {"dikelola sendiri agrinas", "dikelola langsung agrinas", "satgas", "satgas pkh", "penguasaan agrinas"}:
            actor_edges.append((a, pol, estate, "pengelola_langsung/satgas"))
            continue
        actor_edges.append((a, pol, estate, "penerima_kso"))
        actor_polres[a].add(pol)
        actor_estates[a].add(estate)
    # also tag eks as node
    if estate and estate.lower() not in {"-", "—"}:
        actor_polres[estate].add(pol)
        actor_estates[estate].add(estate)

# Alias merge for known hubs
ALIASES = {
    "Bernas Mulya Mandiri": ["Bernas Mulya Mandiri", "Mitra Personal PT. BERNAS MULYA MANDIRI", "Bernas Mulya Mandiri / Osten Panjaitan"],
    "PT Runggu": ["PT Runggu", "PT Runggu (Sdr. Pakpahan)"],
    "PT Palma Agung Betuah (PAB)": ["PT Palma Agung Betuah (PAB)", "PAB", "PT. PAB"],
    "PT Nusantara Sawit Majuma": ["PT Nusantara Sawit Majuma", "Nusantara Sawit Majuma", "Majuma", "Mazuma"],
    "PT Ujung Tanjung Sejahtera": ["PT Ujung Tanjung Sejahtera", "Ujung Tanjung Sejahtera", "UTS"],
    "PT Riden Jaya Konstruksi": ["PT Riden Jaya Konstruksi", "Riden Jaya", "RIDEN JAYA"],
    "PT Tiga Raja Mas": ["PT Tiga Raja Mas (H.M. Ali)", "Tiga Raja Mas"],
    "Poktan Riau Jaya Makmur": ["Poktan Riau Jaya Makmur", "Poktan Riau Jaya Makmur / Paruh Marta"],
    "Berlian Nusantara Perkasa": ["Berlian Nusantara Perkasa", "Berlian NP"],
    "Maju Serempak": ["Maju Serempak"],
    "Agrinas Palma Nusantara": ["PT Agrinas Palma Nusantara", "Agrinas", "PT Agrinas"],
    "CV Ginting Sukses Abadi": ["CV Ginting Sukses Abadi", "Ginting Sukses Abadi"],
    "Digjaya Nata Persada": ["Digjaya Nata Persada"],
    "PT Permata Kencana Utama": ["PT Permata Kencana Utama", "PKU", "vendor PKU"],
}


def canon(name):
    n = name.strip()
    for c, al in ALIASES.items():
        for a in al:
            if a.lower() in n.lower() or n.lower() in a.lower():
                return c
    return n


hub_stats = defaultdict(lambda: {"polres": set(), "estates": set(), "peran": set()})
for a, pol, estate, peran in actor_edges:
    c = canon(a)
    hub_stats[c]["polres"].add(pol)
    hub_stats[c]["estates"].add(estate)
    hub_stats[c]["peran"].add(peran)

# Also force-add known PAM non-BUJP hubs even if sparse
for pol, names in PAM_NON_BUJP.items():
    for n in names:
        c = canon(n)
        hub_stats[c]["polres"].add(pol)
        hub_stats[c]["peran"].add("pam_or_kso_flagged")

hub_rows = []
for actor, st in hub_stats.items():
    n_pol = len(st["polres"])
    n_est = len([e for e in st["estates"] if e])
    if n_pol < 1 and n_est < 1:
        continue
    # risk hub if multi-polres or multi-estate or flagged PAM
    multi = n_pol >= 2 or n_est >= 2
    pam_flag = "pam" in " ".join(st["peran"]).lower() or actor in {
        "PT Nusantara Sawit Majuma", "PT Palma Agung Betuah (PAB)",
        "PT Ujung Tanjung Sejahtera", "PT Riden Jaya Konstruksi",
    }
    if not (multi or pam_flag or n_est >= 2):
        continue
    hub_rows.append({
        "aktor": actor,
        "jml_polres": n_pol,
        "polres": ", ".join(sorted(st["polres"])),
        "jml_estate_terkait": n_est,
        "contoh_estate": "; ".join(sorted(list(st["estates"]))[:3]),
        "peran": ", ".join(sorted(st["peran"])),
        "flag_hub_risiko": "YA" if (multi or pam_flag) else "TIDAK",
        "catatan": "PAM/KSO non-BUJP terkait bentrok" if pam_flag else ("Multi-lokasi" if multi else ""),
    })

hub_rows.sort(key=lambda x: (-x["jml_polres"], -x["jml_estate_terkait"], x["aktor"]))

# Edge list for network sheet (top hubs only)
edge_rows = []
seen = set()
for a, pol, estate, peran in actor_edges:
    c = canon(a)
    key = (c, pol, estate)
    if key in seen:
        continue
    seen.add(key)
    if any(h["aktor"] == c for h in hub_rows[:40]):
        edge_rows.append({
            "source_aktor": c,
            "target_polres": pol,
            "estate": estate,
            "relasi": peran,
        })

# ===================== R4 TIMELINES =====================
timelines = {
    "T1_Majuma_Rohul": [
        ("2025-akhir / pra-2026", "Penunjukan KSO PT Nusantara Sawit Majuma atas lahan PT BK1/Berkat Satu (Sontang, Bonai Darussalam); PAM swakarsa non-BUJP (Nias)"),
        ("2026-01-12", "Bentrok masyarakat 3 desa + PAM Berkat Satu vs PAM Majuma KSO Agrinas (~260 orang); narasi 8 luka"),
        ("2026-01-21", "LP/B/32/I/2026 — penganiayaan terhadap patroli Agrinas+Satgas di Afd IX"),
        ("2026-02-07", "Serangan PAM Majuma ke barak/PAM KUD Telago Biru / Berkat Satu; **1 MD + 6 luka**; LP/B/07/II/2026 Polsek Bonai"),
        ("2026-02-22", "LP/B/21/II/2026 Tambusai — massa ~400 paksa masuk mess Afd VIII Tambusai Timur (konteks Torganda/Agrinas paralel)"),
        ("2026-02-12..23", "Update Intelkam: daftar tersangka Majuma (jumlah berbeda antar deck 12.02 vs 23.02)"),
        ("2026-03-31", "LP/B/107/III/2026 — pencurian TBS Agrinas eks Torganda Afd XII (lanjutan ketegangan panen)"),
        ("Window kritis", "≈ 4–8 minggu dari gesekan Jan menuju MD Feb; pengamanan non-BUJP = akselerator"),
    ],
    "T2_PAB_SIS_Bengkalis": [
        ("pra-Des 2025", "Lahan PT SIS disita Satgas PKH; KSO PT Palma Agung Betuah (PAB); PAM Nias & Sakai non-BUJP"),
        ("2025-12-03", "Bentrok karyawan SIS vs masyarakat Sakai; 1 warga luka kepala"),
        ("2025-12-22", "Bentrok PAB vs SIS saat PAB masuk lahan sitaan; luka bacok; 13 mobil dirusak; LP 476/XII/2025 & LP/B/153/XII/2025"),
        ("2026-01-14", "Warga Bukit Abas vs security PAB → bakar Pos 1; 2 LP; 5 tersangka"),
        ("2026-04..05", "Manajemen PAB berjalan; ketegangan dengan pok H. Rusman/Risman Tobing; panen liar ~150 Ha"),
        ("2026-05-15/16", "Penganiayaan/pembakaran motor; LP/219/V/2026 (PAB) & LP/B/59/V/2026 (kubu Tobing)"),
        ("2026-07-28/29", "Kasus paralel Mutiara Naga–PKU: mediasi gagal → bentrok (sinyal pola sama di Bengkalis)"),
        ("Window kritis", "Eskalasi berulang 6+ bulan di lokasi sama; mediasi tanpa penyelesaian legitimasi kelola gagal meredam"),
    ],
    "T3_UTS_Rohil": [
        ("pra-Okt 2025", "KSO PT Ujung Tanjung Sejahtera atas Ex Rumbia I / PT Gunung Mas Raya (~1.800 Ha); PAM Flores non-BUJP"),
        ("2025-10-20", "Kelompok W. Siringo-Ringo panen sepihak Blok 29–30 vs security UTS (air cabai); **7 luka**; RJ"),
        ("pasca RJ", "Klaster tetap MERAH di Intelkam Feb 2026; penolakan KSO di kebun kuning lain Rohil berlanjut (Ivomas, APSL, Rama Salomo)"),
        ("Gap", "3 LP Agrinas Rohil disebut agregat Bismillah — nomor belum teritemisasi di sumber workspace"),
        ("Window kritis", "Eskalasi cepat ke kekerasan fisik pada episode panen; RJ menutup perkara tetapi tidak menghapus klaster merah"),
    ],
    "T4_DMMP_Dumai": [
        ("pra-2026", "PT Duta Mas Makmur Perkasa sitaan PKH ±1.458,7 Ha (split ~1000 Ha Bengkalis + ~458,7 Ha Dumai)"),
        ("penunjukan KSO", "KSO PT Riden Jaya Konstruksi"),
        ("pra-20 Feb 2026", "Aksi penghadangan karyawan DMMP terhadap karyawan Riden Jaya"),
        ("2026-02-20", "Riden Jaya turun lapangan take-over lahan sitaan; bentrok/tumpang tindih kelola"),
        ("Intelkam Feb 2026", "Masuk klaster KUNING bersama SRPO & Pelintung Jaya (perebutan pekerjaan)"),
        ("Window kritis", "Take-over fisik tanpa konsensus tenaga kerja lama → bentrok; risiko double-count BKS–Dumai"),
    ],
}

# ===================== R5 / R7 GAP FILL =====================
# Pelalawan 9 KSO from Intelkam reconstruction (plan said 10; Intelkam cards = 9 companies)
pelalawan_kso = [
    ("PKSO-01", "Mitra Unggul Pusaka", "Kop. Karya Indragiri Maju", 366.8, 304.01, "Segati/Langgam", "Hijau", "Intelkam 12.02", "Perlu LP satker"),
    ("PKSO-02", "Sari Lembah Subur", "Maju Serempak", 681.01, 19, "Genduang", "Hijau", "Intelkam 12.02", "Perlu LP satker"),
    ("PKSO-03", "Mekar Sari Alam Lestari", "Berlian Nusantara Perkasa", 4745, None, "Mak Teduh", "Hijau", "Intelkam 12.02", "Perlu LP satker"),
    ("PKSO-04", "Serikat Putra", "Agrinas", 138, None, "Sialang Godang", "Hijau", "Intelkam 12.02", "Perlu LP satker"),
    ("PKSO-05", "Gandaerah Hendana (bagian Pelalawan)", "Berlian Nusantara Perkasa", 10, 10, "Kerumutan", "Hijau", "Intelkam 12.02", "Sebagian luas di Inhu"),
    ("PKSO-06", "Eka Sari Lorena", "Kop. Segati Sejahtera", 127, None, "Segati", "Hijau", "Intelkam 12.02", "Perlu LP satker"),
    ("PKSO-07", "Guna Dodos", "Maju Serempak", 5, None, "Sei Kijang", "Hijau", "Intelkam 12.02", "Perlu LP satker"),
    ("PKSO-08", "Agrita Sari Prima", "Kop. Desa Segati", 111.82, None, "Segati", "Hijau", "Intelkam 12.02", "Perlu LP satker"),
    ("PKSO-09", "Viktorindo Alam Lestari", "Poktan LMMB", 85, None, "Palas, PKL Kuras", "Hijau", "Intelkam 12.02", "Perlu LP satker"),
    ("PKSO-10", "Agregat lokal '10 KSO' (nama ke-10 belum di Intelkam cards)", "BELUM TERIDENTIFIKASI", 5392.83, 5392.83, "Kab. Pelalawan", "—", "Pelalawan PDF agregat", "Minta Polres/Disbun isi nama KSO ke-10"),
]

# Rohil LP gap register — what we know + placeholders to fill
rohil_lp_gap = [
    ("RGAP-01", "Bentrok UTS–Siringo 20 Okt 2025", "RJ (nomor LP tidak di Intelkam)", "2025-10-20", "7 luka", "TERCATAT_TANPA_NOMOR", "Minta Polres Rohil kirim nomor LP/RJ resmi"),
    ("RGAP-02", "LP Agrinas #1 (agregat Bismillah=3)", "BELUM ADA", "—", "—", "KOSONG", "Itemisasi wajib"),
    ("RGAP-03", "LP Agrinas #2 (agregat Bismillah=3)", "BELUM ADA", "—", "—", "KOSONG", "Itemisasi wajib"),
    ("RGAP-04", "LP Agrinas #3 (agregat Bismillah=3)", "BELUM ADA", "—", "—", "KOSONG", "Itemisasi wajib"),
    ("RGAP-05", "Daftar 8 KSO Rohil (hanya hitungan)", "NAMA BELUM LENGKAP", "—", "UTS + Digjaya + Satahi + K21 + Parit Nantinggi + Batang Kumu + Rumpun Sejahtera + ?", "PARSIAL", "Lengkapi 8 nama resmi dari Agrinas/Polres"),
    ("RGAP-06", "Pengaduan masyarakat 194 (agregat)", "TIDAK DIITEMISASI", "—", "Bukan semua Agrinas", "AGGREGAT_SAJA", "Sample 20 pengaduan terkait KSO/Agrinas"),
]

# Templates Meranti & Siak
template_rows = [
    ("TMPL-MER-01", "Kep. Meranti", "kebun", "eks_perusahaan", "WAJIB", "Nama badan usaha / perorangan sitaan PKH"),
    ("TMPL-MER-02", "Kep. Meranti", "kebun", "penerima_kso", "WAJIB", "Nama KSO atau 'Belum KSO' / 'Langsung Agrinas'"),
    ("TMPL-MER-03", "Kep. Meranti", "kebun", "luas_sita_ha / luas_kso_ha", "WAJIB", "Angka Disbun/Satgas; satuan Ha"),
    ("TMPL-MER-04", "Kep. Meranti", "kebun", "lokasi (kec/desa)", "WAJIB", ""),
    ("TMPL-MER-05", "Kep. Meranti", "kebun", "klaster + situasi terkini", "WAJIB", "Merah/Kuning/Hijau + 2 kalimat situasi"),
    ("TMPL-MER-06", "Kep. Meranti", "kasus", "nomor_lp / pengaduan", "WAJIB_JIKA_ADA", "Isi NIHIL jika benar nihil — jangan dikosongkan diam-diam"),
    ("TMPL-MER-07", "Kep. Meranti", "kasus", "kategori / pihak / dampak", "WAJIB_JIKA_ADA", ""),
    ("TMPL-MER-08", "Kep. Meranti", "meta", "tanggal update satker", "WAJIB", "Bulanan"),
    ("TMPL-SIAK-01", "Siak", "kebun", "filter: hanya sawit PKH–Agrinas", "WAJIB", "Kecualikan HTI/RAPP/Tahura kecuali overlap sawit sitaan"),
    ("TMPL-SIAK-02", "Siak", "kebun", "luas_keseluruhan / pribadi / KSO / Agrinas", "WAJIB", "Satu baris satu perusahaan; jangan tukar kolom"),
    ("TMPL-SIAK-03", "Siak", "kebun", "penerima_kso + status", "WAJIB", ""),
    ("TMPL-SIAK-04", "Siak", "kasus", "pisahkan Setda-admin vs LP Polres", "WAJIB", "Kolom sumber: Polres / Setda / Disbun"),
    ("TMPL-SIAK-05", "Siak", "kasus", "kait_agrinas_kso (Ya/Tidak/Lemah)", "WAJIB", "Cegah false positive dari sengketa HTI"),
    ("TMPL-SIAK-06", "Siak", "meta", "verifikasi angka Arara 62.193 Ha", "PRIORITAS", "OCR suspect"),
    ("TMPL-ROHIL-01", "Rohil", "kasus", "3 LP Agrinas nomor lengkap", "PRIORITAS", "Lihat RGAP-02..04"),
    ("TMPL-ROHIL-02", "Rohil", "kebun", "8 nama KSO resmi", "PRIORITAS", "Lihat RGAP-05"),
    ("TMPL-PEL-01", "Pelalawan", "kebun", "nama KSO ke-10 + LP Bunut/SSS/MUP", "PRIORITAS", "Lihat PKSO-10"),
]

gap_summary = [
    ("Kep. Meranti", "NIHIL total + file Tebing Tinggi corrupt", "Kirim template TMPL-MER; minta konfirmasi tertulis NIHIL atau isi 1 kebun Intelkam"),
    ("Pekanbaru", "Agrinas NIHIL jelas; 7 LP pencurian non-Agrinas", "Pertahankan status NIHIL Agrinas; jangan campur ke skor KSO"),
    ("Rohil", "24 kebun / 8 KSO / 3 LP Agrinas tanpa nomor", "Itemisasi LP + lengkapi nama KSO"),
    ("Pelalawan", "9/10 KSO teridentifikasi dari Intelkam; LP tipis; TNTN terpisah", "Isi KSO#10 + LP bernomor; pisah register TNTN"),
    ("Siak", "Kolom luas kacau; HTI tercampur", "Template filter sawit PKH; verifikasi Ha"),
]

# ===================== WRITE EXCEL =====================
wb = Workbook()

# R1 satker
ws = wb.active
ws.title = "R1_skor_satker"
headers = list(satker_scores[0].keys())
ws.append(headers)
for row in satker_scores:
    ws.append([row[h] for h in headers])
style_header(ws, len(headers))
for i, row in enumerate(satker_scores, start=2):
    band = row["band_risiko"]
    fill = {"KRITIS": red_fill, "TINGGI": orange_fill, "SEDANG": yellow_fill, "RENDAH": green_fill}[band]
    ws.cell(i, headers.index("band_risiko") + 1).fill = fill
autosize(ws)

ws = wb.create_sheet("R1_skor_kebun")
headers = list(kebun_scores[0].keys())
ws.append(headers)
for row in kebun_scores:
    ws.append([row[h] for h in headers])
style_header(ws, len(headers))
for i, row in enumerate(kebun_scores, start=2):
    band = row["band_risiko"]
    fill = {"KRITIS": red_fill, "TINGGI": orange_fill, "SEDANG": yellow_fill, "RENDAH": green_fill}[band]
    ws.cell(i, headers.index("band_risiko") + 1).fill = fill
autosize(ws)

ws = wb.create_sheet("R1_metodologi")
ws.append(["komponen", "bobot_max", "keterangan"])
for r in [
    ("skor_klaster", 30, "Rata-rata tertimbang Merah30/Kuning18/Hijau4 dari hitungan Intelkam"),
    ("skor_sinyal", 25, "Dari field sinyal_konflik ringkasan_polres"),
    ("skor_korban", 20, "MD/luka/kerusakan material terpetakan (cap 20; MD Rohul=25→cap)"),
    ("skor_pam_non_bujp", 15, "15 jika satker punya PAM non-BUJP named; else proxy volume"),
    ("skor_kelengkapan_data", 5, "0.5 × skor kualitas (kuat10..nihil2)"),
    ("skor_volume", 8, "0.25 × jml kebun, cap 8"),
    ("normalisasi", "≈/1.08 cap 100", "Band: KRITIS≥70; TINGGI≥55; SEDANG≥35; RENDAH<35"),
    ("kebun_score", "rule-based", "Klaster + kata kunci konflik note + PAM + Belum KSO; cap 100"),
]:
    ws.append(list(r))
style_header(ws, 3)
autosize(ws)

# R3
ws = wb.create_sheet("R3_validasi_klaster")
vh = ["id", "polres", "estate", "klaster_resmi", "bukti_aktual", "klaster_usulan", "status", "alasan"]
ws.append(vh)
for v in validasi:
    ws.append(list(v))
style_header(ws, len(vh))
autosize(ws)

ws = wb.create_sheet("R3_rekap_status")
ws.append(["status_validasi", "jumlah", "arti"])
arti = {
    "VALID": "Klaster resmi selaras bukti",
    "VALID_DENGAN_CATATAN": "Selaras dengan syarat/pemisahan tipologi",
    "KOREKSI_NAIK": "Understate — usulan naikkan klaster",
    "KANDIDAT_NAIK": "Belum resmi diubah; pantau/usulkan update",
    "PISAHKAN": "Jangan gabung ke klaster KSO",
    "TIDAK_DAPAT_DIVALIDASI": "Data lokal kosong",
    "DI LUAR OBJEK": "Bukan Agrinas–KSO",
}
for s, n in sorted(status_count.items(), key=lambda x: -x[1]):
    ws.append([s, n, arti.get(s, "")])
style_header(ws, 3)
autosize(ws)

# R2
ws = wb.create_sheet("R2_hub_aktor")
hh = list(hub_rows[0].keys()) if hub_rows else ["aktor"]
ws.append(hh)
for row in hub_rows:
    ws.append([row[h] for h in hh])
style_header(ws, len(hh))
autosize(ws)

ws = wb.create_sheet("R2_edges")
eh = ["source_aktor", "target_polres", "estate", "relasi"]
ws.append(eh)
for row in edge_rows:
    ws.append([row[h] for h in eh])
style_header(ws, len(eh))
autosize(ws)

# R4
ws = wb.create_sheet("R4_timeline")
ws.append(["pilot_id", "urutan", "waktu", "peristiwa"])
for pid, events in timelines.items():
    for i, (t, e) in enumerate(events, 1):
        ws.append([pid, i, t, e])
style_header(ws, 4)
autosize(ws)

# R5 gap
ws = wb.create_sheet("R5_pelalawan_kso")
ws.append(["id", "eks_perusahaan", "penerima_kso", "luas_sita_ha", "luas_kso_ha", "lokasi", "klaster", "sumber", "tindak_lanjut"])
for r in pelalawan_kso:
    ws.append(list(r))
style_header(ws, 9)
autosize(ws)

ws = wb.create_sheet("R5_rohil_lp_gap")
ws.append(["id", "uraian", "nomor_lp", "tanggal", "dampak", "status_data", "tindak_lanjut"])
for r in rohil_lp_gap:
    ws.append(list(r))
style_header(ws, 7)
autosize(ws)

ws = wb.create_sheet("R5_template_isian")
ws.append(["id", "polres", "domain", "field", "kewajiban", "petunjuk"])
for r in template_rows:
    ws.append(list(r))
style_header(ws, 6)
autosize(ws)

ws = wb.create_sheet("R5_gap_summary")
ws.append(["polres", "gap", "aksi"])
for r in gap_summary:
    ws.append(list(r))
style_header(ws, 3)
autosize(ws)

wb.save(OUT_XLSX)
print(f"Saved {OUT_XLSX}")
print("Top satker:", [(s["polres"], s["skor_total"], s["band_risiko"]) for s in satker_scores[:5]])
print("Top kebun:", [(k["polres"], k["skor_kebun"], str(k["eks_perusahaan"])[:40]) for k in kebun_scores[:8]])
print("Validasi status:", dict(status_count))
print("Hubs:", len(hub_rows), "edges:", len(edge_rows))

# ===================== MARKDOWN REPORT =====================
lines = []
lines.append("# Laporan 5 Analisis Lanjutan — Agrinas–KSO Polda Riau")
lines.append("")
lines.append("**Unit:** Unit II Harda · **Tanggal:** 3 Agustus 2026  ")
lines.append(f"**Workbook:** [`analisis_lanjutan_5_prioritas.xlsx`](analisis_lanjutan_5_prioritas.xlsx)  ")
lines.append("**Dasar:** `matriks_agrinas_kso_12_polres.xlsx` + Intelkam/satker  ")
lines.append("")
lines.append("---")
lines.append("")
lines.append("## R1. Matriks skor risiko satker & kebun")
lines.append("")
lines.append("### Metodologi singkat")
lines.append("Skor satker (0–100) = klaster tertimbang + sinyal konflik + korban + PAM non-BUJP + kelengkapan data + volume kebun. Band: **KRITIS ≥70 · TINGGI ≥55 · SEDANG ≥35 · RENDAH <35**.")
lines.append("")
lines.append("### Ranking satker")
lines.append("")
lines.append("| Rank | Polres | Skor | Band | Klaster | Sinyal | Korban | PAM |")
lines.append("|---:|---|---:|---|---:|---:|---:|---:|")
for i, s in enumerate(satker_scores, 1):
    lines.append(f"| {i} | {s['polres']} | {s['skor_total']} | {s['band_risiko']} | {s['skor_klaster']} | {s['skor_sinyal']} | {s['skor_korban']} | {s['skor_pam_non_bujp']} |")
lines.append("")
lines.append("### 10 kebun skor tertinggi")
lines.append("")
lines.append("| Rank | Polres | Skor | Band | Estate | Klaster |")
lines.append("|---:|---|---:|---|---|---|")
for i, k in enumerate(kebun_scores[:10], 1):
    lines.append(f"| {i} | {k['polres']} | {k['skor_kebun']} | {k['band_risiko']} | {str(k['eks_perusahaan'])[:55]} | {k['klaster_intelkam']} |")
lines.append("")
lines.append("**Implikasi R1:** Koridor utara (Rohul–Rohil–Bengkalis) mendominasi band KRITIS/TINGGI. Inhu/Kuansing/Kampar masuk TINGGI karena volume atau LP, meski tanpa MD. Pelalawan naik karena faktor TNTN pada skor korban/aksi — tetap pisahkan analisis KSO vs TNTN saat operasionalisasi.")
lines.append("")
lines.append("---")
lines.append("")
lines.append("## R3. Validasi klaster Intelkam vs kejadian aktual")
lines.append("")
lines.append(f"Diuji **{len(validasi)}** titik estate/portofolio. Rekap status:")
lines.append("")
lines.append("| Status | Jml | Arti |")
lines.append("|---|---:|---|")
for s, n in sorted(status_count.items(), key=lambda x: -x[1]):
    lines.append(f"| {s} | {n} | {arti.get(s, '')} |")
lines.append("")
lines.append("### Yang perlu dikoreksi / dinaikkan")
lines.append("")
for v in validasi:
    if v[6] in {"KOREKSI_NAIK", "KANDIDAT_NAIK"}:
        lines.append(f"- **{v[0]} {v[1]} — {v[2]}:** resmi `{v[3]}` → usulan `{v[5]}` — {v[7]}")
lines.append("")
lines.append("### Yang harus dipisahkan dari klaster KSO")
lines.append("")
for v in validasi:
    if v[6] == "PISAHKAN":
        lines.append(f"- **{v[2]}** ({v[1]}): {v[7]}")
lines.append("")
lines.append("**Implikasi R3:** Definisi merah untuk 3 titik resmi **valid**. Understatement utama ada di Bengkalis (kuning tersembunyi), Inhu EX Palm, Kuansing WJT (dimensi struktural), Kampar Johan Sentosa, dan Mutiara Naga (pasca-briefing).")
lines.append("")
lines.append("---")
lines.append("")
lines.append("## R2. Jaringan aktor KSO / PAM multi-lokasi")
lines.append("")
lines.append(f"Teridentifikasi **{len(hub_rows)} hub aktor** (multi-polres, multi-estate, atau flag PAM/KSO risiko). Cuplikan hub utama:")
lines.append("")
lines.append("| Aktor | #Polres | Polres | #Estate | Flag | Catatan |")
lines.append("|---|---:|---|---:|---|---|")
for h in hub_rows[:20]:
    lines.append(f"| {h['aktor'][:50]} | {h['jml_polres']} | {h['polres']} | {h['jml_estate_terkait']} | {h['flag_hub_risiko']} | {h['catatan']} |")
lines.append("")
lines.append("### Pola jaringan")
lines.append("")
lines.append("```mermaid")
lines.append("flowchart LR")
lines.append("  Agrinas[Agrinas_APN] --> KSO[Penerima_KSO]")
lines.append("  KSO --> PAM[PAM_Swakarsa]")
lines.append("  KSO --> Kebun[Kebun_Sitaan_PKH]")
lines.append("  PAM --> Bentrok[Bentrok_LP]")
lines.append("  Eks[Eks_Penguasa_Karyawan] --> Bentrok")
lines.append("  Masy[Masyarakat_Poktan_MHA] --> Bentrok")
lines.append("```")
lines.append("")
lines.append("**Hub risiko prioritas pantau:** Majuma (Rohul), PAB (Bengkalis), UTS (Rohil), Riden Jaya (Dumai/BKS), Bernas Mulya Mandiri (multi-estate Inhu), Berlian NP & Maju Serempak (multi-estate Pelalawan), Poktan Riau Jaya Makmur (multi Kampar).")
lines.append("")
lines.append("**Implikasi R2:** Risiko tidak hanya “lokasi”, tetapi **aktor yang berpindah antar kebun**. Penerima multi-estate tanpa rekam mediasi lokal memperbesar peluang penolakan berulang.")
lines.append("")
lines.append("---")
lines.append("")
lines.append("## R4. Timeline eskalasi (4 pilot)")
lines.append("")
for pid, events in timelines.items():
    title = {
        "T1_Majuma_Rohul": "Pilot 1 — Majuma / Berkat Satu (Rohul) → MD",
        "T2_PAB_SIS_Bengkalis": "Pilot 2 — PAB–SIS (Bengkalis) → eskalasi berulang",
        "T3_UTS_Rohil": "Pilot 3 — UTS / Gunung Mas Raya (Rohil) → 7 luka",
        "T4_DMMP_Dumai": "Pilot 4 — DMMP–Riden Jaya (Dumai) → take-over",
    }[pid]
    lines.append(f"### {title}")
    lines.append("")
    for t, e in events:
        lines.append(f"- **{t}:** {e}")
    lines.append("")
lines.append("### Pola lead-time lintas pilot")
lines.append("")
lines.append("| Pilot | Pemicu awal | Waktu ke kekerasan signifikan | Akselerator |")
lines.append("|---|---|---|---|")
lines.append("| Majuma | Penolakan KSO + PAM tandingan | ≈ 3–4 minggu (Jan→7 Feb) | PAM non-BUJP, mobilisasi massal |")
lines.append("| PAB–SIS | Masuk lahan sitaan vs eks-karyawan | Hari/minggu; berulang 6+ bulan | Dual LP, PAM, klaim hak kelola |")
lines.append("| UTS | Panen sepihak vs security | Episode tunggal cepat (20 Okt) | PAM non-BUJP, sengketa panen |")
lines.append("| DMMP | Penghadangan → take-over | Menjelang aksi 20 Feb | Split wilayah BKS–Dumai, tenaga kerja lama |")
lines.append("")
lines.append("**Implikasi R4:** Window pengawasan paling kritis adalah **30 hari setelah penunjukan KSO / first entry** ke lahan sitaan, terutama jika PAM non-BUJP dan eks-karyawan masih menguasai operasional.")
lines.append("")
lines.append("---")
lines.append("")
lines.append("## R5. Gap-fill data satker tipis")
lines.append("")
lines.append("### Pelalawan — daftar KSO (9 teridentifikasi + 1 gap)")
lines.append("")
lines.append("| ID | Eks perusahaan | Penerima KSO | Sita Ha | Lokasi | Status data |")
lines.append("|---|---|---|---:|---|---|")
for r in pelalawan_kso:
    lines.append(f"| {r[0]} | {r[1]} | {r[2]} | {r[3] if r[3] is not None else '—'} | {r[5]} | {r[8]} |")
lines.append("")
lines.append("### Rohil — register gap LP / KSO")
lines.append("")
for r in rohil_lp_gap:
    lines.append(f"- **{r[0]}** [{r[5]}]: {r[1]} — tindak lanjut: {r[6]}")
lines.append("")
lines.append("### Ringkasan gap & aksi")
lines.append("")
lines.append("| Polres | Gap | Aksi |")
lines.append("|---|---|---|")
for r in gap_summary:
    lines.append(f"| {r[0]} | {r[1]} | {r[2]} |")
lines.append("")
lines.append("Template isian detail ada di sheet `R5_template_isian` (Meranti, Siak, Rohil, Pelalawan).")
lines.append("")
lines.append("---")
lines.append("")
lines.append("## Sintesis lintas 5 analisis")
lines.append("")
lines.append("1. **Skor risiko** dan **validasi klaster** saling mengunci: Rohul/Rohil/Bengkalis kritis; sejumlah hijau perlu naik.")
lines.append("2. **Jaringan aktor** menjelaskan mengapa eskalasi menular antar kebun (PAM/KSO yang sama atau pola penunjukan serupa).")
lines.append("3. **Timeline** memberi window 30 hari sebagai fokus monitoring analitik pasca-entry KSO.")
lines.append("4. **Gap-fill** adalah prasyarat agar skor R1 untuk Rohil/Pelalawan/Siak/Meranti tidak bias karena data hilang.")
lines.append("")
lines.append("### Prioritas kerja analitik berikutnya (urut)")
lines.append("")
lines.append("1. Update klaster untuk item KOREKSI_NAIK / KANDIDAT_NAIK (R3)")
lines.append("2. Isi RGAP Rohil + PKSO-10 Pelalawan (R5)")
lines.append("3. Perdalam network Majuma–PAB–UTS–Riden + Bernas Mulya (R2)")
lines.append("4. Pantau window 30 hari untuk setiap penunjukan KSO baru (R4)")
lines.append("5. Refresh skor R1 bulanan setelah gap terisi")
lines.append("")
lines.append("*Dokumen analitik Unit II Harda — bukan rekomendasi taktis operasi.*")
lines.append("")

with open(OUT_MD, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print(f"Saved {OUT_MD}")
